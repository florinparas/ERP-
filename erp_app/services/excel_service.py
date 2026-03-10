"""
Serviciu de integrare cu Microsoft Excel via openpyxl.
Export/Import date ERP in format .xlsx compatibil cu Excel.
"""
import os
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from erp_app import db
from erp_app.models import Client, Product, Invoice, Order, Employee, InventoryMovement


HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def export_clients():
    """Exporta lista de clienti in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clienti"

    ws.cell(row=1, column=1, value="Raport Clienti ERP").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["ID", "Nume", "Email", "Telefon", "Companie", "CUI", "Adresa", "Oras", "Tara"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    for row, client in enumerate(Client.query.order_by(Client.name).all(), 5):
        ws.cell(row=row, column=1, value=client.id)
        ws.cell(row=row, column=2, value=client.name)
        ws.cell(row=row, column=3, value=client.email)
        ws.cell(row=row, column=4, value=client.phone)
        ws.cell(row=row, column=5, value=client.company)
        ws.cell(row=row, column=6, value=client.cui)
        ws.cell(row=row, column=7, value=client.address)
        ws.cell(row=row, column=8, value=client.city)
        ws.cell(row=row, column=9, value=client.country)

    _auto_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_products():
    """Exporta lista de produse in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Produse"

    ws.cell(row=1, column=1, value="Catalog Produse ERP").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["Cod", "Nume", "Categorie", "UM", "Pret", "Cost", "TVA%", "Stoc", "Stoc Min", "Activ"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    for row, p in enumerate(Product.query.order_by(Product.name).all(), 5):
        ws.cell(row=row, column=1, value=p.code)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.category)
        ws.cell(row=row, column=4, value=p.unit)
        ws.cell(row=row, column=5, value=p.price).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=p.cost).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=p.vat_rate)
        ws.cell(row=row, column=8, value=p.stock_quantity)
        ws.cell(row=row, column=9, value=p.min_stock)
        ws.cell(row=row, column=10, value="Da" if p.is_active else "Nu")

    _auto_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_invoices():
    """Exporta facturile in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturi"

    ws.cell(row=1, column=1, value="Raport Facturi ERP").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["Nr. Factura", "Client", "Data", "Scadenta", "Status", "Subtotal", "TVA", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    for row, inv in enumerate(Invoice.query.order_by(Invoice.date.desc()).all(), 5):
        ws.cell(row=row, column=1, value=inv.number)
        ws.cell(row=row, column=2, value=inv.client.name if inv.client else "")
        ws.cell(row=row, column=3, value=inv.date.strftime("%d.%m.%Y") if inv.date else "")
        ws.cell(row=row, column=4, value=inv.due_date.strftime("%d.%m.%Y") if inv.due_date else "")
        ws.cell(row=row, column=5, value=inv.status)
        ws.cell(row=row, column=6, value=inv.subtotal).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=inv.vat_total).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=inv.total).number_format = '#,##0.00'

    _auto_width(ws)

    # Add summary sheet
    ws2 = wb.create_sheet("Sumar")
    ws2.cell(row=1, column=1, value="Sumar Facturi").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Total facturi:")
    ws2.cell(row=3, column=2, value=Invoice.query.count())
    ws2.cell(row=4, column=1, value="Platite:")
    ws2.cell(row=4, column=2, value=Invoice.query.filter_by(status="paid").count())
    ws2.cell(row=5, column=1, value="Neplatite:")
    ws2.cell(row=5, column=2, value=Invoice.query.filter_by(status="sent").count())
    paid_total = db.session.query(db.func.coalesce(db.func.sum(Invoice.total), 0)).filter_by(status="paid").scalar()
    ws2.cell(row=6, column=1, value="Valoare incasata:")
    ws2.cell(row=6, column=2, value=paid_total).number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_orders():
    """Exporta comenzile in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comenzi"

    ws.cell(row=1, column=1, value="Raport Comenzi ERP").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["Nr. Comanda", "Client", "Data", "Livrare", "Status", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    for row, order in enumerate(Order.query.order_by(Order.date.desc()).all(), 5):
        ws.cell(row=row, column=1, value=order.number)
        ws.cell(row=row, column=2, value=order.client.name if order.client else "")
        ws.cell(row=row, column=3, value=order.date.strftime("%d.%m.%Y") if order.date else "")
        ws.cell(row=row, column=4, value=order.delivery_date.strftime("%d.%m.%Y") if order.delivery_date else "")
        ws.cell(row=row, column=5, value=order.status)
        ws.cell(row=row, column=6, value=order.total).number_format = '#,##0.00'

    _auto_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_employees():
    """Exporta lista de angajati in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Angajati"

    ws.cell(row=1, column=1, value="Raport Angajati ERP").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["Nume", "Prenume", "Email", "Telefon", "Pozitie", "Departament", "Salariu", "Data Angajare", "Activ"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    for row, emp in enumerate(Employee.query.order_by(Employee.last_name).all(), 5):
        ws.cell(row=row, column=1, value=emp.last_name)
        ws.cell(row=row, column=2, value=emp.first_name)
        ws.cell(row=row, column=3, value=emp.email)
        ws.cell(row=row, column=4, value=emp.phone)
        ws.cell(row=row, column=5, value=emp.position)
        ws.cell(row=row, column=6, value=emp.department)
        ws.cell(row=row, column=7, value=emp.salary).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=emp.hire_date.strftime("%d.%m.%Y") if emp.hire_date else "")
        ws.cell(row=row, column=9, value="Da" if emp.is_active else "Nu")

    _auto_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_inventory():
    """Exporta stocul curent in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stoc"

    ws.cell(row=1, column=1, value="Situatia Stocurilor").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    headers = ["Cod", "Produs", "Categorie", "UM", "Stoc Curent", "Stoc Minim", "Pret", "Valoare Stoc"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row, p in enumerate(Product.query.order_by(Product.name).all(), 5):
        ws.cell(row=row, column=1, value=p.code)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.category)
        ws.cell(row=row, column=4, value=p.unit)
        stock_cell = ws.cell(row=row, column=5, value=p.stock_quantity)
        ws.cell(row=row, column=6, value=p.min_stock)
        ws.cell(row=row, column=7, value=p.price).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=p.stock_quantity * p.price).number_format = '#,##0.00'
        if p.stock_quantity <= p.min_stock:
            stock_cell.fill = alert_fill

    _auto_width(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_clients(file_data):
    """Importa clienti din fisier Excel."""
    wb = load_workbook(file_data)
    ws = wb.active
    imported = 0
    errors = []

    for row in range(5, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        try:
            client = Client(
                name=str(name),
                email=str(ws.cell(row=row, column=3).value or ""),
                phone=str(ws.cell(row=row, column=4).value or ""),
                company=str(ws.cell(row=row, column=5).value or ""),
                cui=str(ws.cell(row=row, column=6).value or ""),
                address=str(ws.cell(row=row, column=7).value or ""),
                city=str(ws.cell(row=row, column=8).value or ""),
                country=str(ws.cell(row=row, column=9).value or "Romania"),
            )
            db.session.add(client)
            imported += 1
        except Exception as e:
            errors.append(f"Rand {row}: {str(e)}")

    if imported > 0:
        db.session.commit()
    return imported, errors


def import_products(file_data):
    """Importa produse din fisier Excel."""
    wb = load_workbook(file_data)
    ws = wb.active
    imported = 0
    errors = []

    for row in range(5, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        if not code or not name:
            continue
        try:
            existing = Product.query.filter_by(code=str(code)).first()
            if existing:
                existing.name = str(name)
                existing.category = str(ws.cell(row=row, column=3).value or "")
                existing.unit = str(ws.cell(row=row, column=4).value or "buc")
                existing.price = float(ws.cell(row=row, column=5).value or 0)
                existing.cost = float(ws.cell(row=row, column=6).value or 0)
                existing.vat_rate = float(ws.cell(row=row, column=7).value or 19)
            else:
                product = Product(
                    code=str(code),
                    name=str(name),
                    category=str(ws.cell(row=row, column=3).value or ""),
                    unit=str(ws.cell(row=row, column=4).value or "buc"),
                    price=float(ws.cell(row=row, column=5).value or 0),
                    cost=float(ws.cell(row=row, column=6).value or 0),
                    vat_rate=float(ws.cell(row=row, column=7).value or 19),
                    stock_quantity=float(ws.cell(row=row, column=8).value or 0),
                    min_stock=float(ws.cell(row=row, column=9).value or 0),
                )
                db.session.add(product)
            imported += 1
        except Exception as e:
            errors.append(f"Rand {row}: {str(e)}")

    if imported > 0:
        db.session.commit()
    return imported, errors
