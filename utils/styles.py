"""
Stiluri Excel reutilizabile pentru ERP HR Module
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from config.hr_config import COLORS


def get_header_font():
    return Font(name="Calibri", size=11, bold=True, color=COLORS["header_font"])


def get_header_fill():
    return PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")


def get_title_font():
    return Font(name="Calibri", size=14, bold=True, color=COLORS["title_bg"])


def get_title_fill():
    return PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")


def get_title_font_white():
    return Font(name="Calibri", size=14, bold=True, color=COLORS["header_font"])


def get_alt_row_fill():
    return PatternFill(start_color=COLORS["row_alt"], end_color=COLORS["row_alt"], fill_type="solid")


def get_normal_row_fill():
    return PatternFill(start_color=COLORS["row_normal"], end_color=COLORS["row_normal"], fill_type="solid")


def get_border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def get_center_alignment():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def get_left_alignment():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def get_status_fill(status_type):
    """Returnează fill-ul pentru diferite statusuri"""
    fills = {
        "success": PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid"),
        "warning": PatternFill(start_color=COLORS["light_yellow"], end_color=COLORS["light_yellow"], fill_type="solid"),
        "danger": PatternFill(start_color=COLORS["light_red"], end_color=COLORS["light_red"], fill_type="solid"),
        "info": PatternFill(start_color=COLORS["light_blue"], end_color=COLORS["light_blue"], fill_type="solid"),
        "orange": PatternFill(start_color=COLORS["light_orange"], end_color=COLORS["light_orange"], fill_type="solid"),
    }
    return fills.get(status_type, get_normal_row_fill())


def apply_header_style(ws, row, num_cols):
    """Aplică stilul header pe un rând"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()


def apply_data_style(ws, row, num_cols, is_alt=False):
    """Aplică stilul pe un rând de date"""
    fill = get_alt_row_fill() if is_alt else get_normal_row_fill()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = get_left_alignment()
        cell.border = get_border()
        cell.font = Font(name="Calibri", size=10)


def apply_sheet_title(ws, title, num_cols):
    """Adaugă un titlu de foaie pe primul rând (merge cells)"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = get_title_font_white()
    cell.fill = get_title_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws, min_width=10, max_width=40):
    """Ajustează automat lățimea coloanelor"""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def setup_print_area(ws, num_cols, num_rows):
    """Configurare print area"""
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(num_cols)
    ws.print_area = f"A1:{last_col}{num_rows}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
