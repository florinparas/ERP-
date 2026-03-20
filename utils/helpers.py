"""
Funcții helper pentru generarea Excel-ului ERP HR
"""
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from config.hr_config import COLORS
from utils.styles import (
    apply_header_style, apply_data_style, apply_sheet_title,
    auto_fit_columns, get_border, get_center_alignment
)


def create_sheet_with_headers(wb, sheet_name, title, headers, data=None,
                               header_row=2, freeze_row=3):
    """
    Creează o foaie cu titlu, headere și date opționale.

    Args:
        wb: Workbook
        sheet_name: Numele foii
        title: Titlul afișat
        headers: Lista de tuple (nume_coloana, latime)
        data: Lista de liste cu date (opțional)
        header_row: Rândul pe care se pune header-ul
        freeze_row: Rândul de la care se freeze

    Returns:
        ws: Worksheet-ul creat
    """
    ws = wb.create_sheet(title=sheet_name)
    num_cols = len(headers)

    # Titlu
    apply_sheet_title(ws, title, num_cols)

    # Headere
    for col_idx, (col_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    apply_header_style(ws, header_row, num_cols)

    # Date
    if data:
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            is_alt = (row_idx - header_row) % 2 == 0
            apply_data_style(ws, row_idx, num_cols, is_alt)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # Auto filter
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

    return ws


def add_dropdown_validation(ws, cell_range, options, allow_blank=True):
    """
    Adaugă validare dropdown pe un range de celule.

    Args:
        ws: Worksheet
        cell_range: Range (ex: "K3:K1000")
        options: Lista de opțiuni sau formula referință
        allow_blank: Permite celule goale
    """
    if isinstance(options, list):
        formula = '"' + ','.join(str(o) for o in options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    else:
        # Referință la un named range sau foaie
        dv = DataValidation(type="list", formula1=options, allow_blank=allow_blank)

    dv.error = "Valoare invalidă! Selectați din lista disponibilă."
    dv.errorTitle = "Eroare Validare"
    dv.prompt = "Selectați o opțiune din listă"
    dv.promptTitle = "Selectare"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_number_validation(ws, cell_range, min_val=None, max_val=None):
    """Adaugă validare numerică pe un range"""
    dv = DataValidation(type="whole", operator="between",
                        formula1=str(min_val) if min_val is not None else "0",
                        formula2=str(max_val) if max_val is not None else "999999999")
    dv.error = f"Introduceți un număr între {min_val} și {max_val}."
    dv.errorTitle = "Eroare Validare"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_status_conditional_formatting(ws, col_letter, start_row, end_row,
                                        status_colors):
    """
    Adaugă formatare condițională bazată pe text.

    Args:
        ws: Worksheet
        col_letter: Litera coloanei (ex: "O")
        start_row: Rândul de start
        end_row: Rândul de final
        status_colors: Dict {text: fill_color_hex}
    """
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for text, color in status_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{text}"'],
            fill=fill,
            font=Font(name="Calibri", size=10, bold=True)
        )
        ws.conditional_formatting.add(cell_range, rule)


def add_score_conditional_formatting(ws, col_letter, start_row, end_row):
    """Adaugă formatare condițională gradient pentru scoruri 1-5"""
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

    score_colors = {
        1: "FF0000",  # Roșu
        2: "FF8C00",  # Portocaliu
        3: "FFD700",  # Galben
        4: "90EE90",  # Verde deschis
        5: "008000",  # Verde
    }

    for score, color in score_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font_color = "FFFFFF" if score in (1, 5) else "000000"
        rule = CellIsRule(
            operator="equal",
            formula=[str(score)],
            fill=fill,
            font=Font(name="Calibri", size=10, bold=True, color=font_color)
        )
        ws.conditional_formatting.add(cell_range, rule)


def add_date_expiry_formatting(ws, col_letter, start_row, end_row, days_warning=30):
    """Evidențiază datele care expiră în N zile"""
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

    # Expirat (roșu)
    rule_expired = FormulaRule(
        formula=[f'AND({col_letter}{start_row}<>"",{col_letter}{start_row}<TODAY())'],
        fill=PatternFill(start_color=COLORS["light_red"], end_color=COLORS["light_red"], fill_type="solid"),
        font=Font(color="FF0000", bold=True)
    )
    ws.conditional_formatting.add(cell_range, rule_expired)

    # Expiră curând (galben)
    rule_warning = FormulaRule(
        formula=[f'AND({col_letter}{start_row}<>"",{col_letter}{start_row}>=TODAY(),{col_letter}{start_row}<=TODAY()+{days_warning})'],
        fill=PatternFill(start_color=COLORS["light_yellow"], end_color=COLORS["light_yellow"], fill_type="solid"),
        font=Font(color="FF8C00", bold=True)
    )
    ws.conditional_formatting.add(cell_range, rule_warning)


def vlookup_formula(lookup_cell, table_sheet, table_col_start, table_col_end,
                     col_index, match_type=0):
    """
    Generează o formulă VLOOKUP.

    Args:
        lookup_cell: Celula de căutare (ex: "B3")
        table_sheet: Numele foii tabel
        table_col_start: Coloana start (ex: "A")
        table_col_end: Coloana final (ex: "C")
        col_index: Indexul coloanei de returnat
        match_type: 0 = exact match
    """
    return (f'=IFERROR(VLOOKUP({lookup_cell},'
            f"'{table_sheet}'!${table_col_start}:${table_col_end},"
            f'{col_index},{match_type}),"N/A")')


def add_pontaj_conditional_formatting(ws, start_col, end_col, start_row, end_row):
    """Adaugă formatare condițională pentru codurile de pontaj"""
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    cell_range = f"{start_letter}{start_row}:{end_letter}{end_row}"

    pontaj_colors = {
        "P": COLORS["light_green"],
        "CO": COLORS["light_blue"],
        "CM": COLORS["light_orange"],
        "A": COLORS["light_red"],
        "AM": COLORS["light_yellow"],
        "LS": "E0E0E0",
        "OS": "E1BEE7",
        "TP": "B2DFDB",
        "D": "FFE0B2",
    }

    for code, color in pontaj_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{code}"'],
            fill=fill,
            font=Font(name="Calibri", size=9, bold=True)
        )
        ws.conditional_formatting.add(cell_range, rule)
