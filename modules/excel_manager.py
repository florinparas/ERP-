import os
from openpyxl import Workbook, load_workbook


def init_file(file_path, columns):
    """Create an Excel file with headers if it doesn't exist."""
    if os.path.exists(file_path):
        return
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Date'
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    wb.save(file_path)


def _get_headers(ws):
    """Read header row and return list of column names."""
    return [cell.value for cell in ws[1]]


def _row_to_dict(headers, row):
    """Convert a worksheet row to a dictionary."""
    return {headers[i]: row[i].value for i in range(len(headers))}


def read_all(file_path):
    """Read all records from an Excel file."""
    if not os.path.exists(file_path):
        return []
    wb = load_workbook(file_path)
    ws = wb.active
    headers = _get_headers(ws)
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        records.append(_row_to_dict(headers, row))
    wb.close()
    return records


def read_by_id(file_path, record_id):
    """Read a single record by ID."""
    record_id = int(record_id)
    if not os.path.exists(file_path):
        return None
    wb = load_workbook(file_path)
    ws = wb.active
    headers = _get_headers(ws)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == record_id:
            result = _row_to_dict(headers, row)
            wb.close()
            return result
    wb.close()
    return None


def create(file_path, data):
    """Create a new record. Returns the new ID."""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = _get_headers(ws)

    # Find next ID
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value is not None and isinstance(row[0].value, (int, float)):
            max_id = max(max_id, int(row[0].value))
    new_id = max_id + 1
    data['id'] = new_id

    # Append row
    next_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=next_row, column=col_idx, value=data.get(header))

    wb.save(file_path)
    wb.close()
    return new_id


def update(file_path, record_id, data):
    """Update a record by ID."""
    record_id = int(record_id)
    wb = load_workbook(file_path)
    ws = wb.active
    headers = _get_headers(ws)

    for row in ws.iter_rows(min_row=2):
        if row[0].value == record_id:
            for col_idx, header in enumerate(headers, 1):
                if header in data and header != 'id':
                    ws.cell(row=row[0].row, column=col_idx, value=data[header])
            break

    wb.save(file_path)
    wb.close()


def delete(file_path, record_id):
    """Delete a record by ID."""
    record_id = int(record_id)
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == record_id:
            ws.delete_rows(row[0].row)
            break

    wb.save(file_path)
    wb.close()


def filter_by(file_path, **kwargs):
    """Filter records by column values."""
    records = read_all(file_path)
    results = []
    for record in records:
        match = True
        for key, value in kwargs.items():
            if str(record.get(key, '')) != str(value):
                match = False
                break
        if match:
            results.append(record)
    return results


def count(file_path):
    """Count total records."""
    return len(read_all(file_path))
