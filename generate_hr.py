#!/usr/bin/env python3
"""
ERP Management Personal - Generator Excel
Generează fișierul Excel cu modulul HR complet (16 foi).

Utilizare:
    python generate_hr.py [--output FILENAME]
"""
import argparse
import sys

try:
    from openpyxl import Workbook
except ImportError:
    print("EROARE: openpyxl nu este instalat.")
    print("Rulați: pip install openpyxl")
    sys.exit(1)

from sheets import (
    configurare,
    departamente,
    angajati,
    contracte,
    documente,
    pontaj,
    ore_suplimentare,
    concedii,
    salarizare,
    fluturasi,
    evaluari,
    training,
    recrutare,
    organigrama,
    istoric,
    dashboard,
)
from config.hr_config import COLORS


def generate_hr_module(output_path="ERP_HR_Module.xlsx"):
    """Generează fișierul Excel cu modulul HR complet."""
    print("=" * 60)
    print("  ERP Management Personal - Generator Excel")
    print("=" * 60)

    wb = Workbook()
    wb.remove(wb.active)

    # Ordinea foilor: Dashboard primul, apoi în ordine logică
    # Configurare și Departamente trebuie create devreme (referite de celelalte)
    modules = [
        ("Dashboard", dashboard),
        ("Angajați", angajati),
        ("Contracte", contracte),
        ("Departamente", departamente),
        ("Documente", documente),
        ("Pontaj", pontaj),
        ("Ore Suplimentare", ore_suplimentare),
        ("Concedii", concedii),
        ("Salarizare", salarizare),
        ("Fluturași", fluturasi),
        ("Evaluări", evaluari),
        ("Training", training),
        ("Recrutare", recrutare),
        ("Organigramă", organigrama),
        ("Istoric", istoric),
        ("Configurare", configurare),
    ]

    for sheet_name, module in modules:
        print(f"  → Generare foaie: {sheet_name}...")
        try:
            module.create_sheet(wb)
        except Exception as e:
            print(f"    EROARE la {sheet_name}: {e}")
            raise

    # Culori tab-uri
    tab_colors = {
        "Dashboard": COLORS["title_bg"],
        "Angajați": COLORS["accent"],
        "Contracte": COLORS["info"],
        "Departamente": COLORS["success"],
        "Documente": "795548",
        "Pontaj": "FF8C00",
        "Ore Suplimentare": "E65100",
        "Concedii": COLORS["info"],
        "Salarizare": "70AD47",
        "Fluturași": "43A047",
        "Evaluări": "9C27B0",
        "Training": "FF9800",
        "Recrutare": "00BCD4",
        "Organigramă": "3F51B5",
        "Istoric": "757575",
        "Configurare": "607D8B",
    }

    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    # Dashboard ca foaie activă
    wb.active = wb.sheetnames.index("Dashboard")

    # Salvare
    wb.save(output_path)
    print()
    print(f"  Fișier generat cu succes: {output_path}")
    print(f"  Foi create: {len(wb.sheetnames)}")
    print(f"  Foi: {', '.join(wb.sheetnames)}")
    print("=" * 60)

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generator ERP Modul HR - Excel"
    )
    parser.add_argument(
        "--output", "-o",
        default="ERP_HR_Module.xlsx",
        help="Numele fișierului Excel de output (default: ERP_HR_Module.xlsx)"
    )
    args = parser.parse_args()
    generate_hr_module(args.output)


if __name__ == "__main__":
    main()
