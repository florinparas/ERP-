#!/usr/bin/env python3
"""
StockAgent — Generator Portofoliu Investiții Excel
Generează fișierul Excel cu modulul complet de management portofoliu.

Utilizare:
    python generate_stock_portfolio.py [--output FILENAME]
"""
import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("EROARE: openpyxl nu este instalat.")
    print("Rulați: pip install openpyxl")
    sys.exit(1)

from sheets.stock import (
    configurare,
    pozitii,
    tranzactii,
    analiza_tehnica,
    analiza_fundamentala,
    watchlist,
    risk_management,
    dashboard,
)
from config.stock_config import TAB_COLORS


def generate_stock_portfolio(output_path="StockAgent_Portfolio.xlsx"):
    """Generează fișierul Excel cu modulul StockAgent complet."""
    print("=" * 60)
    print("  StockAgent — Generator Portofoliu Investiții")
    print("  Broker Virtual | 20 ani experiență | Profil Agresiv")
    print("=" * 60)

    wb = Workbook()
    # Ștergem foaia default
    wb.remove(wb.active)

    # Creăm foile în ordinea corectă
    modules = [
        ("Dashboard", dashboard),
        ("Poziții", pozitii),
        ("Tranzacții", tranzactii),
        ("Analiză Tehnică", analiza_tehnica),
        ("Analiză Fundamentală", analiza_fundamentala),
        ("Watchlist", watchlist),
        ("Risk Management", risk_management),
        ("Configurare", configurare),
    ]

    for sheet_name, module in modules:
        print(f"  → Generare foaie: {sheet_name}...")
        try:
            module.create_sheet(wb)
        except Exception as e:
            print(f"    EROARE la {sheet_name}: {e}")
            raise

    # Setăm culori tab-uri
    for sheet_name, color in TAB_COLORS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    # Dashboard ca foaie activă
    wb.active = wb.sheetnames.index("Dashboard")

    # Salvare
    wb.save(output_path)
    print()
    print(f"  ✓ Fișier generat cu succes: {output_path}")
    print(f"  ✓ Foi create: {len(wb.sheetnames)}")
    print(f"  ✓ Foi: {', '.join(wb.sheetnames)}")
    print("=" * 60)

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="StockAgent — Generator Portofoliu Investiții Excel"
    )
    parser.add_argument(
        "--output", "-o",
        default="StockAgent_Portfolio.xlsx",
        help="Numele fișierului Excel de output (default: StockAgent_Portfolio.xlsx)"
    )
    args = parser.parse_args()
    generate_stock_portfolio(args.output)


if __name__ == "__main__":
    main()
