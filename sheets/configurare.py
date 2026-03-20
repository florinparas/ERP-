"""
Foaia Configurare - Setări sistem, parametri fiscali, liste dropdown
"""
from openpyxl.utils import get_column_letter
from config.hr_config import (
    CAS_RATE, CASS_RATE, TAX_RATE, CAM_RATE, SALARIU_MINIM_BRUT,
    DEDUCERE_PERSONALA_BAZA, ZILE_CONCEDIU_STANDARD, ORE_LUCRU_ZI,
    SARBATORI_LEGALE, STATUS_ANGAJAT, TIPURI_CONTRACT, TIPURI_CONCEDIU,
    STATUS_CONCEDIU, CODURI_PONTAJ, SEXE, NIVEL_FUNCTIE, STATUS_RECRUTARE,
    SURSE_RECRUTARE, TIPURI_TRAINING, STATUS_TRAINING,
    NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_RON, COLORS
)
from utils.styles import (
    apply_sheet_title, get_header_font, get_header_fill, get_border,
    get_center_alignment, get_left_alignment, get_alt_row_fill,
    auto_fit_columns
)


def create_sheet(wb):
    """Creează foaia Configurare"""
    ws = wb.create_sheet(title="Configurare")

    # ---- SECȚIUNEA 1: PARAMETRI FISCALI ----
    _section_header(ws, 1, "PARAMETRI FISCALI ROMÂNIA 2025", 3)

    params = [
        ("Parametru", "Valoare", "Descriere"),
        ("CAS (Pensie)", CAS_RATE, "Contribuție asigurări sociale - angajat"),
        ("CASS (Sănătate)", CASS_RATE, "Contribuție sănătate - angajat"),
        ("Impozit pe Venit", TAX_RATE, "Impozit pe venit - flat tax"),
        ("CAM", CAM_RATE, "Contribuția asiguratorie muncă - angajator"),
        ("Salariu Minim Brut", SALARIU_MINIM_BRUT, "Salariul minim brut pe economie (RON)"),
        ("Deducere Personală Bază", DEDUCERE_PERSONALA_BAZA, "Deducere personală de bază (RON)"),
        ("Zile CO Standard", ZILE_CONCEDIU_STANDARD, "Zile concediu odihnă / an"),
        ("Ore Lucru / Zi", ORE_LUCRU_ZI, "Program normal de lucru"),
    ]

    for row_idx, (param, val, desc) in enumerate(params, 2):
        ws.cell(row=row_idx, column=1, value=param)
        cell_val = ws.cell(row=row_idx, column=2, value=val)
        ws.cell(row=row_idx, column=3, value=desc)

        if row_idx == 2:
            # Header row
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = get_header_font()
                cell.fill = get_header_fill()
                cell.alignment = get_center_alignment()
                cell.border = get_border()
        else:
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = get_border()
                cell.alignment = get_left_alignment()
            # Formatare valoare
            if isinstance(val, float) and val < 1:
                cell_val.number_format = NUMBER_FORMAT_PERCENT
            elif isinstance(val, (int, float)) and val >= 100:
                cell_val.number_format = NUMBER_FORMAT_RON

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45

    # ---- SECȚIUNEA 2: DATE COMPANIE ----
    company_start = len(params) + 3
    _section_header(ws, company_start, "DATE COMPANIE", 3)

    company_fields = [
        ("Denumire Companie", "SC Example SRL"),
        ("CUI", "RO12345678"),
        ("Nr. Reg. Comerțului", "J40/1234/2010"),
        ("Adresă Sediu", "Str. Exemplu 10, București, Sector 1"),
        ("Telefon", "+40 21 123 4567"),
        ("Email", "office@example.ro"),
        ("IBAN", "RO49AAAA1B31007593840000"),
        ("Bancă", "Banca Transilvania"),
    ]

    for row_idx, (field, value) in enumerate(company_fields, company_start + 1):
        ws.cell(row=row_idx, column=1, value=field).border = get_border()
        ws.cell(row=row_idx, column=2, value=value).border = get_border()

    # ---- SECȚIUNEA 3: LISTE DROPDOWN (coloanele E-L) ----
    lists_col_start = 5  # Coloana E
    dropdown_lists = [
        ("Status Angajat", STATUS_ANGAJAT),
        ("Tipuri Contract", TIPURI_CONTRACT),
        ("Tipuri Concediu", TIPURI_CONCEDIU),
        ("Status Concediu", STATUS_CONCEDIU),
        ("Coduri Pontaj", CODURI_PONTAJ),
        ("Sex", SEXE),
        ("Nivel Funcție", NIVEL_FUNCTIE),
        ("Status Recrutare", STATUS_RECRUTARE),
        ("Surse Recrutare", SURSE_RECRUTARE),
        ("Tipuri Training", TIPURI_TRAINING),
        ("Status Training", STATUS_TRAINING),
    ]

    # Titlu secțiune
    _section_header(ws, 1, "LISTE DROPDOWN", len(dropdown_lists),
                    start_col=lists_col_start)

    for col_offset, (list_name, items) in enumerate(dropdown_lists):
        col = lists_col_start + col_offset
        # Header
        header_cell = ws.cell(row=2, column=col, value=list_name)
        header_cell.font = get_header_font()
        header_cell.fill = get_header_fill()
        header_cell.alignment = get_center_alignment()
        header_cell.border = get_border()

        # Items
        for row_idx, item in enumerate(items, 3):
            cell = ws.cell(row=row_idx, column=col, value=item)
            cell.border = get_border()
            cell.alignment = get_left_alignment()

        # Lățime coloană
        ws.column_dimensions[get_column_letter(col)].width = max(
            len(list_name) + 2,
            max(len(str(i)) for i in items) + 2
        )

    # ---- SECȚIUNEA 4: SĂRBĂTORI LEGALE ----
    holidays_col = lists_col_start + len(dropdown_lists) + 1
    for year, dates in SARBATORI_LEGALE.items():
        header_cell = ws.cell(row=2, column=holidays_col, value=f"Sărbători {year}")
        header_cell.font = get_header_font()
        header_cell.fill = get_header_fill()
        header_cell.alignment = get_center_alignment()
        header_cell.border = get_border()

        for row_idx, date_str in enumerate(dates, 3):
            cell = ws.cell(row=row_idx, column=holidays_col, value=date_str)
            cell.border = get_border()
            cell.alignment = get_center_alignment()

        ws.column_dimensions[get_column_letter(holidays_col)].width = 18
        holidays_col += 1

    # Freeze panes
    ws.freeze_panes = "A3"

    return ws


def _section_header(ws, row, title, num_cols, start_col=1):
    """Adaugă un header de secțiune"""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row, end_column=start_col + num_cols - 1
    )
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["header_font"])
    cell.fill = PatternFill(
        start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
