"""
Foaia Organigramă - Vizualizare ierarhie organizațională
"""
from config.hr_config import ANGAJATI_DEMO, COLORS, NUMBER_FORMAT_INT
from utils.styles import (
    apply_sheet_title, get_header_font, get_header_fill, get_border,
    get_center_alignment, get_left_alignment
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet(wb):
    """Creează foaia Organigramă cu date ierarhice și vizualizare text"""
    ws = wb.create_sheet(title="Organigramă")

    # ============================================================
    # TABEL 1: DATE IERARHICE (coloanele A-G)
    # ============================================================
    data_headers = [
        ("ID Angajat", 12),
        ("Nume Prenume", 22),
        ("Funcție", 24),
        ("Departament", 20),
        ("Manager Direct (ID)", 16),
        ("Manager Direct (Nume)", 22),
        ("Nivel Ierarhic", 14),
    ]

    apply_sheet_title(ws, "ORGANIGRAMĂ COMPANIE", len(data_headers))

    for col_idx, (name, width) in enumerate(data_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Date auto din foaia Angajați (formule)
    for row in range(3, 53):  # Suport până la 50 angajați
        r = row
        # ID Angajat
        ws.cell(row=r, column=1).value = (
            f"=IF('Angajați'!A{r}<>\"\",\'Angajați\'!A{r},\"\")"
        )
        ws.cell(row=r, column=1).border = get_border()

        # Nume Prenume
        ws.cell(row=r, column=2).value = (
            f"=IF(A{r}<>\"\",\'Angajați\'!B{r}&\" \"&\'Angajați\'!C{r},\"\")"
        )
        ws.cell(row=r, column=2).border = get_border()

        # Funcție
        ws.cell(row=r, column=3).value = (
            f"=IF(A{r}<>\"\",\'Angajați\'!L{r},\"\")"
        )
        ws.cell(row=r, column=3).border = get_border()

        # Departament
        ws.cell(row=r, column=4).value = (
            f"=IF(A{r}<>\"\",\'Angajați\'!K{r},\"\")"
        )
        ws.cell(row=r, column=4).border = get_border()

        # Manager Direct (ID) - căutăm marca managerului
        ws.cell(row=r, column=5).value = (
            f'=IF(A{r}<>"",IFERROR(INDEX(\'Angajați\'!$A:$A,'
            f'MATCH(\'Angajați\'!M{r},\'Angajați\'!$B:$B&" "&\'Angajați\'!$C:$C,0)),'
            f'"-"),"")'
        )
        ws.cell(row=r, column=5).border = get_border()

        # Manager Direct (Nume)
        ws.cell(row=r, column=6).value = (
            f"=IF(A{r}<>\"\",\'Angajați\'!M{r},\"\")"
        )
        ws.cell(row=r, column=6).border = get_border()

        # Nivel Ierarhic (simplificat)
        ws.cell(row=r, column=7).value = (
            f'=IF(A{r}="","",IF(F{r}="-",1,'
            f'IF(IFERROR(MATCH(F{r},B$3:B$52,0),0)>0,2,1)))'
        )
        ws.cell(row=r, column=7).border = get_border()

    # ============================================================
    # TABEL 2: VIZUALIZARE TEXT ORGANIGRAMĂ (coloana I+)
    # ============================================================
    viz_col = 9  # Coloana I
    ws.merge_cells(start_row=1, start_column=viz_col,
                   end_row=1, end_column=viz_col + 4)
    viz_title = ws.cell(row=1, column=viz_col,
                        value="VIZUALIZARE ORGANIGRAMĂ (TEXT)")
    from utils.styles import get_title_font_white, get_title_fill
    viz_title.font = get_title_font_white()
    viz_title.fill = get_title_fill()
    viz_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions[get_column_letter(viz_col)].width = 50
    ws.column_dimensions[get_column_letter(viz_col + 1)].width = 20
    ws.column_dimensions[get_column_letter(viz_col + 2)].width = 20

    # Header vizualizare
    viz_headers = ["Ierarhie", "Funcție", "Departament"]
    for i, h in enumerate(viz_headers):
        cell = ws.cell(row=2, column=viz_col + i, value=h)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()

    # Organigramă statică demo (tree view text)
    org_tree = [
        (0, "Popescu Ion", "Director General", "Management"),
        (1, "├── Ionescu Maria", "Team Lead Dezvoltare", "IT & Dezvoltare"),
        (1, "├── Georgescu Ana", "Specialist HR", "Resurse Umane"),
        (1, "├── Dumitrescu Pavel", "Contabil", "Financiar-Contabil"),
        (1, "├── Marinescu Elena", "Manager Vânzări", "Vânzări"),
        (1, "├── Constantinescu Dan", "Specialist Marketing", "Marketing"),
        (1, "└── Vasilescu Mihai", "Operator Logistică", "Logistică"),
    ]

    level_colors = {
        0: "1F4E79",   # Director - albastru închis
        1: "2E75B6",   # Manager - albastru
        2: "5B9BD5",   # Lead - albastru deschis
        3: "BDD7EE",   # Staff - albastru foarte deschis
    }

    for idx, (level, name, func, dept) in enumerate(org_tree, 3):
        # Ierarhie cu indentare
        indent = "    " * level
        cell_name = ws.cell(row=idx, column=viz_col, value=f"{indent}{name}")
        cell_name.font = Font(
            name="Consolas", size=11,
            bold=(level == 0),
            color="FFFFFF" if level == 0 else "000000"
        )
        color = level_colors.get(level, level_colors[3])
        cell_name.fill = PatternFill(start_color=color, end_color=color,
                                      fill_type="solid")
        cell_name.border = get_border()

        cell_func = ws.cell(row=idx, column=viz_col + 1, value=func)
        cell_func.font = Font(name="Calibri", size=10)
        cell_func.border = get_border()

        cell_dept = ws.cell(row=idx, column=viz_col + 2, value=dept)
        cell_dept.font = Font(name="Calibri", size=10)
        cell_dept.border = get_border()

    # Notă
    note_row = 3 + len(org_tree) + 1
    ws.cell(row=note_row, column=viz_col,
            value="NOTĂ: Organigrama vizuală se actualizează manual sau prin VBA."
            ).font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws.cell(row=note_row + 1, column=viz_col,
            value="Datele din tabelul din stânga se actualizează automat din foaia Angajați."
            ).font = Font(name="Calibri", size=9, italic=True, color="666666")

    # Freeze panes
    ws.freeze_panes = "A3"

    return ws
