"""
Foaia Dashboard - Panou principal HR cu sumarizări și grafice
"""
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.hr_config import COLORS, NUMBER_FORMAT_RON, NUMBER_FORMAT_INT


def create_sheet(wb):
    """Creează foaia Dashboard"""
    ws = wb.create_sheet(title="Dashboard")

    # ============================================================
    # TITLU PRINCIPAL
    # ============================================================
    ws.merge_cells("A1:L2")
    title_cell = ws.cell(row=1, column=1, value="ERP - MODUL HR | DASHBOARD MANAGEMENT PERSONAL")
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                   end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================
    # KPI-URI (rândul 4-6)
    # ============================================================
    kpi_configs = [
        (1, "Total Angajați", "=COUNTA('Angajați'!A3:A1000)", COLORS["accent"]),
        (3, "Angajați Activi", '=COUNTIF(\'Angajați\'!O3:O1000,"Activ")', COLORS["success"]),
        (5, "Concedii Active",
         '=COUNTIFS(Concedii!H3:H1000,"Aprobat",Concedii!E3:E1000,"<="&TODAY(),Concedii!F3:F1000,">="&TODAY())',
         COLORS["info"]),
        (7, "Posturi Deschise",
         '=COUNTIFS(Recrutare!J3:J1000,"<>Angajat",Recrutare!J3:J1000,"<>Respins",Recrutare!J3:J1000,"<>Retras",Recrutare!J3:J1000,"<>")',
         COLORS["warning"]),
        (9, "Fond Salarii Lunar",
         "=SUM(Salarizare!V3:V100)", "2E75B6"),
        (11, "Cost Total Angajator",
         "=SUM(Salarizare!X3:X100)", COLORS["danger"]),
    ]

    for start_col, label, formula, color in kpi_configs:
        _create_kpi_card(ws, start_col, 4, label, formula, color)

    # ============================================================
    # SECȚIUNEA STATISTICI (rândul 8+)
    # ============================================================
    # Secțiune: Angajați pe departament (pentru grafic)
    ws.cell(row=8, column=1, value="ANGAJAȚI PE DEPARTAMENT").font = Font(
        name="Calibri", size=12, bold=True, color=COLORS["title_bg"]
    )

    # Date pentru grafic (vor fi calculate cu formule)
    dept_data_start = 9
    ws.cell(row=dept_data_start, column=1, value="Departament").font = Font(bold=True)
    ws.cell(row=dept_data_start, column=2, value="Nr. Angajați").font = Font(bold=True)

    for i in range(7):  # Max 7 departamente din demo
        row = dept_data_start + 1 + i
        ws.cell(row=row, column=1).value = f"=IF(Departamente!B{3+i}<>\"\",Departamente!B{3+i},\"\")"
        ws.cell(row=row, column=2).value = f"=IF(A{row}<>\"\",Departamente!F{3+i},0)"

    # Grafic Pie - Angajați pe departament
    pie = PieChart()
    pie.title = "Distribuție Angajați pe Departament"
    pie.style = 10
    data = Reference(ws, min_col=2, min_row=dept_data_start,
                     max_row=dept_data_start + 7)
    cats = Reference(ws, min_col=1, min_row=dept_data_start + 1,
                     max_row=dept_data_start + 7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.width = 18
    pie.height = 12
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    ws.add_chart(pie, "D8")

    # ============================================================
    # SECȚIUNEA ALERTE (coloana A, rândul 20+)
    # ============================================================
    ws.merge_cells("A20:C20")
    alert_header = ws.cell(row=20, column=1, value="ALERTE & NOTIFICĂRI")
    alert_header.font = Font(name="Calibri", size=12, bold=True,
                              color=COLORS["header_font"])
    alert_header.fill = PatternFill(start_color=COLORS["danger"],
                                     end_color=COLORS["danger"], fill_type="solid")

    alerts = [
        ("Contracte care expiră în 30 zile:",
         '=COUNTIFS(Contracte!G3:G1000,">="&TODAY(),Contracte!G3:G1000,"<="&TODAY()+30)'),
        ("Evaluări scadente în 30 zile:",
         '=COUNTIFS(\'Evaluări\'!N3:N1000,">="&TODAY(),\'Evaluări\'!N3:N1000,"<="&TODAY()+30)'),
        ("Documente expirate:",
         '=COUNTIF(Documente!I3:I1000,"Expirat")'),
        ("Certificări care expiră în 90 zile:",
         '=COUNTIFS(Training!M3:M1000,">="&TODAY(),Training!M3:M1000,"<="&TODAY()+90)'),
        ("Concedii în așteptare:",
         '=COUNTIF(Concedii!H3:H1000,"În Așteptare")'),
        ("Ore suplimentare nesoluționate:",
         '=COUNTIF(\'Ore Suplimentare\'!L3:L1000,"Solicitat")'),
        ("Candidați în proces recrutare:",
         '=COUNTIFS(Recrutare!J3:J1000,"<>Angajat",Recrutare!J3:J1000,"<>Respins",Recrutare!J3:J1000,"<>Retras",Recrutare!J3:J1000,"<>")'),
    ]

    for idx, (alert_label, formula) in enumerate(alerts, 21):
        ws.cell(row=idx, column=1, value=alert_label).font = Font(
            name="Calibri", size=10)
        ws.cell(row=idx, column=1).border = Border(
            bottom=Side(style="thin", color=COLORS["border"]))
        val_cell = ws.cell(row=idx, column=3)
        val_cell.value = formula
        val_cell.font = Font(name="Calibri", size=12, bold=True)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.border = Border(
            bottom=Side(style="thin", color=COLORS["border"]))

    # ============================================================
    # SECȚIUNEA SUMAR SALARIZARE (coloana A, rândul 28+)
    # ============================================================
    ws.merge_cells("A30:C30")
    sal_header = ws.cell(row=30, column=1, value="SUMAR SALARIZARE")
    sal_header.font = Font(name="Calibri", size=12, bold=True,
                            color=COLORS["header_font"])
    sal_header.fill = PatternFill(start_color=COLORS["accent"],
                                   end_color=COLORS["accent"], fill_type="solid")

    sal_items = [
        ("Total Brut (incl. OS, CO):", "=SUM(Salarizare!L3:L100)"),
        ("Total CAS (pensie):", "=SUM(Salarizare!M3:M100)"),
        ("Total CASS (sănătate):", "=SUM(Salarizare!N3:N100)"),
        ("Total Impozit:", "=SUM(Salarizare!R3:R100)"),
        ("Total Rețineri:", "=SUM(Salarizare!U3:U100)"),
        ("Total Salarii Nete:", "=SUM(Salarizare!V3:V100)"),
        ("Total CAM (angajator):", "=SUM(Salarizare!W3:W100)"),
        ("Cost Total Companie:", "=SUM(Salarizare!X3:X100)"),
    ]

    for idx, (label, formula) in enumerate(sal_items, 31):
        ws.cell(row=idx, column=1, value=label).font = Font(name="Calibri", size=10)
        val_cell = ws.cell(row=idx, column=3)
        val_cell.value = formula
        val_cell.number_format = NUMBER_FORMAT_RON
        val_cell.font = Font(name="Calibri", size=10, bold=True)
        val_cell.alignment = Alignment(horizontal="right")

    # ============================================================
    # NAVIGARE RAPIDĂ (coloana K-L, rândul 20+)
    # ============================================================
    ws.merge_cells("K20:L20")
    nav_header = ws.cell(row=20, column=11, value="NAVIGARE RAPIDĂ")
    nav_header.font = Font(name="Calibri", size=12, bold=True,
                            color=COLORS["header_font"])
    nav_header.fill = PatternFill(start_color=COLORS["accent"],
                                   end_color=COLORS["accent"], fill_type="solid")

    nav_items = [
        "Angajați", "Contracte", "Departamente", "Documente",
        "Pontaj", "Ore Suplimentare", "Concedii", "Salarizare",
        "Fluturași", "Evaluări", "Training", "Recrutare",
        "Organigramă", "Istoric", "Configurare",
    ]

    for idx, sheet_name in enumerate(nav_items, 21):
        cell = ws.cell(row=idx, column=11, value=f"→ {sheet_name}")
        cell.font = Font(name="Calibri", size=10, color=COLORS["accent"],
                         underline="single")
        # Adaugă hyperlink intern
        cell.hyperlink = f"#'{sheet_name}'!A1"

    # ============================================================
    # CONFIGURARE COLOANE
    # ============================================================
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18

    # Tab color
    ws.sheet_properties.tabColor = COLORS["accent"]

    return ws


def _create_kpi_card(ws, start_col, start_row, label, formula, color):
    """Creează un card KPI cu label pe rândul de sus și valoare dedesubt"""
    # Rândul 1: Label (merge 2 coloane)
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col + 1)
    label_cell = ws.cell(row=start_row, column=start_col)
    label_cell.value = label
    label_cell.font = Font(name="Calibri", size=9, bold=True,
                            color=COLORS["header_font"])
    label_cell.fill = PatternFill(start_color=color, end_color=color,
                                   fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True)

    # Rândul 2-3: Valoare (merge 2 coloane, 2 rânduri)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col,
                   end_row=start_row + 2, end_column=start_col + 1)
    val_cell = ws.cell(row=start_row + 1, column=start_col)
    val_cell.value = formula
    val_cell.font = Font(name="Calibri", size=16, bold=True, color=color)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Formatare monetară pentru sume
    if "Salarii" in label or "Cost" in label or "Fond" in label:
        val_cell.number_format = NUMBER_FORMAT_RON
