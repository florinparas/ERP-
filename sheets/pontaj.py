"""
Foaia Pontaj - Pontaj lunar
"""
from config.hr_config import CODURI_PONTAJ, COLORS
from utils.helpers import (
    add_dropdown_validation, add_pontaj_conditional_formatting
)
from utils.styles import (
    apply_sheet_title, get_header_font, get_header_fill, get_border,
    get_center_alignment, get_left_alignment, get_alt_row_fill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def create_sheet(wb):
    """Creează foaia Pontaj"""
    ws = wb.create_sheet(title="Pontaj")

    # Coloane fixe + 31 zile + totaluri
    fixed_headers = [
        ("ID Angajat", 12),
        ("Nume", 20),
        ("Departament", 18),
        ("Luna", 8),
        ("An", 8),
    ]
    day_headers = [(str(d), 4) for d in range(1, 32)]
    total_headers = [
        ("Zile Lucrate", 12),
        ("Total CO", 10),
        ("Total CM", 10),
        ("Total Absențe", 12),
        ("Total OS", 10),
    ]

    all_headers = fixed_headers + day_headers + total_headers
    num_cols = len(all_headers)

    # Titlu
    apply_sheet_title(ws, "PONTAJ LUNAR", num_cols)

    # Headere
    for col_idx, (name, width) in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Date demo - 5 angajați, luna martie 2025
    demo_rows = [
        ("A001", None, None, 3, 2025),
        ("A002", None, None, 3, 2025),
        ("A003", None, None, 3, 2025),
        ("A004", None, None, 3, 2025),
        ("A005", None, None, 3, 2025),
    ]

    day_col_start = len(fixed_headers) + 1  # Coloana 6 = ziua 1
    day_col_end = day_col_start + 30  # Coloana 36 = ziua 31
    total_col_start = day_col_end + 1  # Coloana 37

    for row_idx, (emp_id, _, _, luna, an) in enumerate(demo_rows, 3):
        ws.cell(row=row_idx, column=1, value=emp_id).border = get_border()

        # VLOOKUP pentru Nume
        ws.cell(row=row_idx, column=2).value = (
            f'=IFERROR(VLOOKUP(A{row_idx},\'Angajați\'!$A:$C,2,0)&" "'
            f'&VLOOKUP(A{row_idx},\'Angajați\'!$A:$C,3,0),"N/A")'
        )
        ws.cell(row=row_idx, column=2).border = get_border()

        # VLOOKUP pentru Departament
        ws.cell(row=row_idx, column=3).value = (
            f"=IFERROR(VLOOKUP(A{row_idx},'Angajați'!$A:$K,11,0),\"N/A\")"
        )
        ws.cell(row=row_idx, column=3).border = get_border()

        ws.cell(row=row_idx, column=4, value=luna).border = get_border()
        ws.cell(row=row_idx, column=5, value=an).border = get_border()

        # Zile demo - P pentru zile lucrătoare, LS pentru weekend
        # Martie 2025: 1=sâmbătă, 2=duminică, etc.
        march_2025_pattern = [
            "LS", "LS", "P", "P", "P", "P", "P", "LS", "LS",  # 1-9
            "P", "P", "P", "P", "P", "LS", "LS",              # 10-16
            "P", "P", "P", "P", "P", "LS", "LS",              # 17-23
            "P", "P", "P", "P", "P", "LS", "LS",              # 24-30
            "P",                                                 # 31
        ]

        for day in range(31):
            day_col = day_col_start + day
            cell = ws.cell(row=row_idx, column=day_col,
                          value=march_2025_pattern[day])
            cell.alignment = get_center_alignment()
            cell.border = get_border()
            cell.font = Font(name="Calibri", size=8)

        # Formule totaluri
        first_day_col = get_column_letter(day_col_start)
        last_day_col = get_column_letter(day_col_end)
        day_range = f"{first_day_col}{row_idx}:{last_day_col}{row_idx}"

        # Total Zile Lucrate (P + TP)
        ws.cell(row=row_idx, column=total_col_start).value = (
            f'=COUNTIF({day_range},"P")+COUNTIF({day_range},"TP")'
        )
        ws.cell(row=row_idx, column=total_col_start).border = get_border()

        # Total CO
        ws.cell(row=row_idx, column=total_col_start + 1).value = (
            f'=COUNTIF({day_range},"CO")'
        )
        ws.cell(row=row_idx, column=total_col_start + 1).border = get_border()

        # Total CM
        ws.cell(row=row_idx, column=total_col_start + 2).value = (
            f'=COUNTIF({day_range},"CM")'
        )
        ws.cell(row=row_idx, column=total_col_start + 2).border = get_border()

        # Total Absențe (A + AM)
        ws.cell(row=row_idx, column=total_col_start + 3).value = (
            f'=COUNTIF({day_range},"A")+COUNTIF({day_range},"AM")'
        )
        ws.cell(row=row_idx, column=total_col_start + 3).border = get_border()

        # Total OS
        ws.cell(row=row_idx, column=total_col_start + 4).value = (
            f'=COUNTIF({day_range},"OS")'
        )
        ws.cell(row=row_idx, column=total_col_start + 4).border = get_border()

    # Validare dropdown pe zilele de pontaj
    first_day = get_column_letter(day_col_start)
    last_day = get_column_letter(day_col_end)
    add_dropdown_validation(
        ws, f"{first_day}3:{last_day}200", CODURI_PONTAJ
    )

    # Formatare condițională pentru codurile de pontaj
    add_pontaj_conditional_formatting(ws, day_col_start, day_col_end, 3, 200)

    # Freeze panes - coloane fixe + header
    ws.freeze_panes = ws.cell(row=3, column=6)

    return ws
