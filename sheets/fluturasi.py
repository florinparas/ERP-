"""
Foaia Fluturași - Fluturași de salariu (pay slips) print-ready
"""
from config.hr_config import COLORS, NUMBER_FORMAT_RON
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet(wb):
    """Creează foaia Fluturași cu layout print-ready"""
    ws = wb.create_sheet(title="Fluturași")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    thick_border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium")
    )

    header_fill = PatternFill(start_color=COLORS["header_bg"],
                               end_color=COLORS["header_bg"], fill_type="solid")
    light_fill = PatternFill(start_color=COLORS["light_blue"],
                              end_color=COLORS["light_blue"], fill_type="solid")
    total_fill = PatternFill(start_color=COLORS["title_bg"],
                              end_color=COLORS["title_bg"], fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=10)
    value_font = Font(name="Calibri", size=10, bold=True)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    net_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    # Setare lățimi coloane
    col_widths = {"A": 5, "B": 22, "C": 18, "D": 5, "E": 22, "F": 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ============================================================
    # SELECTARE ANGAJAT (control cell)
    # ============================================================
    ws.cell(row=1, column=1, value="Selectați rândul din Salarizare:").font = Font(
        name="Calibri", size=10, bold=True)
    ctrl_cell = ws.cell(row=1, column=3, value=3)
    ctrl_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["danger"])
    ctrl_cell.alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=4,
            value="← Introduceți nr. rândului din foaia Salarizare (3, 4, 5...)").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    ROW = "C1"  # Referință la celula de control

    # ============================================================
    # FLUTURAȘ DE SALARIU (rândurile 3-35)
    # ============================================================
    r = 3  # Start row

    # --- HEADER COMPANIE ---
    ws.merge_cells(f"A{r}:F{r}")
    cell = ws.cell(row=r, column=1)
    cell.value = "=Configurare!B13"  # Denumire Companie
    cell.font = title_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    cell = ws.cell(row=r, column=1)
    cell.value = "=Configurare!B14"  # CUI
    cell.font = Font(name="Calibri", size=10, color="FFFFFF")
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    # --- TITLU ---
    ws.merge_cells(f"A{r}:F{r}")
    cell = ws.cell(row=r, column=1)
    cell.value = f'="FLUTURAȘ DE SALARIU - "&INDIRECT("Salarizare!D"&{ROW})&"/"&INDIRECT("Salarizare!E"&{ROW})'
    cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["title_bg"])
    cell.alignment = Alignment(horizontal="center")
    r += 2  # Spațiu

    # --- DATE ANGAJAT ---
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value="DATE ANGAJAT").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    r += 1

    emp_fields = [
        ("Marcă:", f'=INDIRECT("Salarizare!B"&{ROW})',
         "Nume Prenume:", f'=INDIRECT("Salarizare!C"&{ROW})'),
        ("Departament:", f'=IFERROR(VLOOKUP(INDIRECT("Salarizare!B"&{ROW}),\'Angajați\'!$A:$K,11,0),"N/A")',
         "Funcție:", f'=IFERROR(VLOOKUP(INDIRECT("Salarizare!B"&{ROW}),\'Angajați\'!$A:$L,12,0),"N/A")'),
        ("CNP:", f'=IFERROR(VLOOKUP(INDIRECT("Salarizare!B"&{ROW}),\'Angajați\'!$A:$D,4,0),"N/A")',
         "IBAN:", f'=IFERROR(VLOOKUP(INDIRECT("Salarizare!B"&{ROW}),\'Angajați\'!$A:$S,19,0),"N/A")'),
    ]

    for label1, val1, label2, val2 in emp_fields:
        ws.cell(row=r, column=1, value=label1).font = label_font
        ws.cell(row=r, column=1).fill = light_fill
        ws.cell(row=r, column=1).border = thin_border
        c = ws.cell(row=r, column=2, value=val1)
        c.font = value_font
        c.border = thin_border

        ws.cell(row=r, column=4, value=label2).font = label_font
        ws.cell(row=r, column=4).fill = light_fill
        ws.cell(row=r, column=4).border = thin_border
        # Merge D+E for label
        ws.cell(row=r, column=5).value = val2
        ws.cell(row=r, column=5).font = value_font
        ws.cell(row=r, column=5).border = thin_border
        r += 1

    r += 1  # Spațiu

    # --- VENITURI ---
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(row=r, column=1, value="VENITURI").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(f"D{r}:F{r}")
    ws.cell(row=r, column=4, value="REȚINERI").font = header_font
    ws.cell(row=r, column=4).fill = header_fill
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    r += 1

    # Sub-headere
    for col_start, labels in [(1, ["Descriere", "Suma (RON)"]),
                               (4, ["Descriere", "Suma (RON)"])]:
        for i, label in enumerate(labels):
            c = ws.cell(row=r, column=col_start + i, value=label)
            c.font = Font(name="Calibri", size=9, bold=True)
            c.fill = light_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
    r += 1

    # Venituri (stânga) și Rețineri (dreapta)
    venituri = [
        ("Salariu Brut Bază", f'=INDIRECT("Salarizare!F"&{ROW})'),
        ("Zile Lucrate / Total", f'=INDIRECT("Salarizare!G"&{ROW})&" / "&INDIRECT("Salarizare!H"&{ROW})'),
        ("Salariu Proporțional", f'=INDIRECT("Salarizare!I"&{ROW})'),
        ("Tichete Masă", f'=INDIRECT("Salarizare!P"&{ROW})'),
    ]

    retineri = [
        ("CAS 25% (Pensie)", f'=INDIRECT("Salarizare!J"&{ROW})'),
        ("CASS 10% (Sănătate)", f'=INDIRECT("Salarizare!K"&{ROW})'),
        ("Impozit 10%", f'=INDIRECT("Salarizare!N"&{ROW})'),
        ("Alte Deduceri", f'=INDIRECT("Salarizare!O"&{ROW})'),
    ]

    max_rows = max(len(venituri), len(retineri))
    for i in range(max_rows):
        if i < len(venituri):
            ws.cell(row=r + i, column=1, value=venituri[i][0]).font = label_font
            ws.cell(row=r + i, column=1).border = thin_border
            val = ws.cell(row=r + i, column=2, value=venituri[i][1])
            val.font = value_font
            val.number_format = NUMBER_FORMAT_RON
            val.border = thin_border
            val.alignment = Alignment(horizontal="right")

        if i < len(retineri):
            ws.cell(row=r + i, column=4, value=retineri[i][0]).font = label_font
            ws.cell(row=r + i, column=4).border = thin_border
            val = ws.cell(row=r + i, column=5, value=retineri[i][1])
            val.font = value_font
            val.number_format = NUMBER_FORMAT_RON
            val.border = thin_border
            val.alignment = Alignment(horizontal="right")

    r += max_rows + 1

    # --- TOTALURI ---
    # Total Venituri
    ws.cell(row=r, column=1, value="TOTAL VENITURI BRUTE").font = Font(
        name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).fill = light_fill
    ws.cell(row=r, column=1).border = thick_border
    val = ws.cell(row=r, column=2,
                  value=f'=INDIRECT("Salarizare!I"&{ROW})+INDIRECT("Salarizare!P"&{ROW})')
    val.font = value_font
    val.number_format = NUMBER_FORMAT_RON
    val.border = thick_border
    val.alignment = Alignment(horizontal="right")

    # Total Rețineri
    ws.cell(row=r, column=4, value="TOTAL REȚINERI").font = Font(
        name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=4).fill = light_fill
    ws.cell(row=r, column=4).border = thick_border
    val = ws.cell(row=r, column=5,
                  value=f'=INDIRECT("Salarizare!J"&{ROW})+INDIRECT("Salarizare!K"&{ROW})'
                        f'+INDIRECT("Salarizare!N"&{ROW})+INDIRECT("Salarizare!O"&{ROW})')
    val.font = value_font
    val.number_format = NUMBER_FORMAT_RON
    val.border = thick_border
    val.alignment = Alignment(horizontal="right")

    r += 2

    # --- SALARIU NET ---
    ws.merge_cells(f"A{r}:F{r}")
    cell = ws.cell(row=r, column=1)
    cell.value = f'="SALARIU NET DE PLATĂ: "&TEXT(INDIRECT("Salarizare!Q"&{ROW}),"#,##0.00")&" RON"'
    cell.font = net_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border
    r += 2

    # --- CONTRIBUȚII ANGAJATOR ---
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value="CONTRIBUȚII ANGAJATOR").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    r += 1

    ws.cell(row=r, column=1, value="CAM 2.25%").font = label_font
    ws.cell(row=r, column=1).border = thin_border
    val = ws.cell(row=r, column=2,
                  value=f'=INDIRECT("Salarizare!R"&{ROW})')
    val.font = value_font
    val.number_format = NUMBER_FORMAT_RON
    val.border = thin_border
    r += 1

    ws.cell(row=r, column=1, value="Cost Total Angajator").font = Font(
        name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).fill = light_fill
    ws.cell(row=r, column=1).border = thick_border
    val = ws.cell(row=r, column=2,
                  value=f'=INDIRECT("Salarizare!S"&{ROW})')
    val.font = value_font
    val.number_format = NUMBER_FORMAT_RON
    val.border = thick_border
    r += 2

    # --- SEMNĂTURI ---
    ws.cell(row=r, column=1, value="Întocmit,").font = label_font
    ws.cell(row=r, column=4, value="Luat la cunoștință,").font = label_font
    r += 2
    ws.cell(row=r, column=1, value="_____________________").font = label_font
    ws.cell(row=r, column=4, value="_____________________").font = label_font
    r += 1
    ws.cell(row=r, column=1, value="(Departament HR)").font = Font(
        name="Calibri", size=9, italic=True)
    ws.cell(row=r, column=4, value="(Angajat)").font = Font(
        name="Calibri", size=9, italic=True)

    # --- CONFIGURARE PRINT ---
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    return ws
