"""
Foaia Risk Management — Metrici risc, calculator position sizing, alocare sectorială
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, SECTOARE, POZITII_DEMO,
    MAX_POSITION_PCT, MAX_POSITION_SPECULATIVE_PCT, MAX_SECTOR_PCT,
    MIN_CASH_RESERVE_PCT, STOP_LOSS_DEFAULT_PCT, STOP_LOSS_MAX_PCT,
    TRAILING_STOP_BREAKEVEN_TRIGGER, TRAILING_STOP_TRAIL_TRIGGER,
    TRAILING_STOP_TRAIL_PCT, MIN_RR_RATIO, MIN_RR_SPECULATIVE,
    DRAWDOWN_REDUCE_PCT, DRAWDOWN_EXIT_PCT, MAX_CONSECUTIVE_LOSSES,
    PORTFOLIO_INITIAL_CAPITAL, PORTFOLIO_PEAK_VALUE,
    NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_INT,
    NUMBER_FORMAT_PRICE,
)


def create_sheet(wb):
    """Creează foaia Risk Management"""
    ws = wb.create_sheet(title="Risk Management")

    # Titlu principal
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | RISK MANAGEMENT — DISCIPLINĂ ABSOLUTĂ")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 1: PORTFOLIO RISK OVERVIEW
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 1 — OVERVIEW RISC PORTOFOLIU", 4)
    current_row += 1

    overview_items = [
        ("Metrică", "Valoare", "Limită", "Status"),
    ]
    # Header row
    for c, val in enumerate(overview_items[0], 1):
        cell = ws.cell(row=current_row, column=c, value=val)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center_align()
        cell.border = _border()
    current_row += 1

    # Metrici dinamice (formule referențiază Poziții)
    metrics = [
        ("Valoare Totală Portofoliu",
         "=SUM('Poziții'!L3:L100)+Configurare!B4",
         f"{PORTFOLIO_INITIAL_CAPITAL:,.0f}",
         '=IF(B{r}>0,"OK","ATENȚIE")'),
        ("Valoare Investită",
         "=SUM('Poziții'!K3:K100)",
         "",
         ""),
        ("Cash Disponibil",
         "=Configurare!B4",
         f"Min {MIN_CASH_RESERVE_PCT*100:.0f}%",
         '=IF(B{r}/B{r0}>{min_cash},"OK","SUB LIMITĂ")'.format(
             r="{r}", r0="{r0}", min_cash=MIN_CASH_RESERVE_PCT)),
        ("Cash % din Portofoliu",
         "=IF(B{r0}>0,B{r_prev}/B{r0},0)",
         f">{MIN_CASH_RESERVE_PCT*100:.0f}%",
         '=IF(B{r}>{min_cash},"OK","SUB LIMITĂ")'.format(
             r="{r}", min_cash=MIN_CASH_RESERVE_PCT)),
        ("P&L Total (RON)",
         "=SUM('Poziții'!M3:M100)",
         "",
         '=IF(B{r}>0,"PROFIT","PIERDERE")'),
        ("P&L Total (%)",
         "=IF(SUM('Poziții'!K3:K100)>0,SUM('Poziții'!M3:M100)/SUM('Poziții'!K3:K100),0)",
         "",
         ""),
        ("Drawdown de la Peak",
         f"=IF({PORTFOLIO_PEAK_VALUE}>0,(B{{{0}}}-{PORTFOLIO_PEAK_VALUE})/{PORTFOLIO_PEAK_VALUE},0)",
         f"Max -{DRAWDOWN_REDUCE_PCT*100:.0f}%",
         ""),
        ("Nr. Poziții Deschise",
         "=COUNTA('Poziții'!A3:A100)",
         "",
         ""),
    ]

    r0 = current_row  # row of first metric (valoare totala)
    for idx, (label, formula, limit, status_formula) in enumerate(metrics):
        r = current_row + idx

        ws.cell(row=r, column=1, value=label).border = _border()
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)

        # Replace placeholders
        formula_resolved = formula.replace("{r}", str(r)).replace("{r0}", str(r0)).replace("{r_prev}", str(r - 1))
        val_cell = ws.cell(row=r, column=2)
        val_cell.value = formula_resolved
        val_cell.border = _border()
        val_cell.font = Font(name="Calibri", size=11, bold=True)
        val_cell.alignment = _center_align()

        ws.cell(row=r, column=3, value=limit).border = _border()
        ws.cell(row=r, column=3).alignment = _center_align()

        if status_formula:
            status_resolved = status_formula.replace("{r}", str(r)).replace("{r0}", str(r0))
            ws.cell(row=r, column=4).value = status_resolved
        ws.cell(row=r, column=4).border = _border()
        ws.cell(row=r, column=4).alignment = _center_align()
        ws.cell(row=r, column=4).font = Font(name="Calibri", size=10, bold=True)

        # Formate
        if "%" in label:
            val_cell.number_format = NUMBER_FORMAT_PERCENT
        elif "Nr." in label:
            val_cell.number_format = NUMBER_FORMAT_INT
        else:
            val_cell.number_format = NUMBER_FORMAT_CURRENCY

    current_row += len(metrics) + 2

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 2: POSITION SIZING CALCULATOR
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 2 — CALCULATOR POSITION SIZING", 4)
    current_row += 1

    calc_label_col = 1
    calc_val_col = 2
    calc_desc_col = 3

    calc_inputs = [
        ("Capital Total (RON):", PORTFOLIO_INITIAL_CAPITAL, "Introdu capitalul total"),
        ("Risk per Trade (%):", 0.02, "Riscul maxim pe o tranzacție (2% recomandat)"),
        ("Preț Intrare:", 30.00, "Prețul la care vrei să cumperi"),
        ("Stop-Loss:", 27.90, "Prețul stop-loss planificat"),
        ("Target:", 36.00, "Prețul target"),
    ]

    input_start = current_row
    for idx, (label, val, desc) in enumerate(calc_inputs):
        r = current_row + idx
        ws.cell(row=r, column=1, value=label).border = _border()
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)

        input_cell = ws.cell(row=r, column=2, value=val)
        input_cell.border = _border()
        input_cell.alignment = _center_align()
        input_cell.font = Font(name="Calibri", size=11, bold=True,
                               color=COLORS["accent"])
        # Highlight input cells
        input_cell.fill = PatternFill(start_color=COLORS["light_yellow"],
                                      end_color=COLORS["light_yellow"],
                                      fill_type="solid")

        ws.cell(row=r, column=3, value=desc).border = _border()
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=9,
                                              color="666666")

        if "%" in label:
            input_cell.number_format = NUMBER_FORMAT_PERCENT
        elif "RON" in label:
            input_cell.number_format = NUMBER_FORMAT_CURRENCY
        else:
            input_cell.number_format = NUMBER_FORMAT_PRICE

    current_row += len(calc_inputs) + 1

    # Rezultate calculate
    cap_row = input_start
    risk_row = input_start + 1
    entry_row = input_start + 2
    sl_row = input_start + 3
    target_row = input_start + 4

    calc_results = [
        ("REZULTATE CALCULATOR", "", ""),
        ("Risk Amount (RON):",
         f"=B{cap_row}*B{risk_row}", "Suma riscată pe această tranzacție"),
        ("Risk per Share (RON):",
         f"=ABS(B{entry_row}-B{sl_row})", "Diferența entry - stop loss"),
        ("Nr. Acțiuni (Position Size):",
         f"=IF(B{current_row+2}>0,ROUNDDOWN(B{current_row+1}/B{current_row+2},0),0)",
         "Calculat: Risk Amount / Risk per Share"),
        ("Valoare Poziție (RON):",
         f"=B{current_row+3}*B{entry_row}", "Valoarea totală a poziției"),
        ("% din Portofoliu:",
         f"=IF(B{cap_row}>0,B{current_row+4}/B{cap_row},0)",
         f"Max {MAX_POSITION_PCT*100:.0f}% standard / {MAX_POSITION_SPECULATIVE_PCT*100:.0f}% speculativ"),
        ("R/R Ratio:",
         f"=IF(B{current_row+2}>0,ABS(B{target_row}-B{entry_row})/B{current_row+2},0)",
         f"Minim {MIN_RR_RATIO:.1f}:1"),
        ("Profit Potențial (RON):",
         f"=B{current_row+3}*(B{target_row}-B{entry_row})",
         "Dacă se atinge targetul"),
        ("Pierdere Maximă (RON):",
         f"=B{current_row+3}*(B{entry_row}-B{sl_row})",
         "Dacă se activează stop-loss"),
    ]

    for idx, (label, formula, desc) in enumerate(calc_results):
        r = current_row + idx
        ws.cell(row=r, column=1, value=label).border = _border()

        if idx == 0:
            # Sub-header
            ws.cell(row=r, column=1).font = Font(name="Calibri", size=11,
                                                  bold=True, color=COLORS["accent"])
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color=COLORS["light_blue"],
                    end_color=COLORS["light_blue"], fill_type="solid")
                ws.cell(row=r, column=c).border = _border()
            continue

        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)
        val_cell = ws.cell(row=r, column=2)
        val_cell.value = formula
        val_cell.border = _border()
        val_cell.alignment = _center_align()
        val_cell.font = Font(name="Calibri", size=11, bold=True)

        ws.cell(row=r, column=3, value=desc).border = _border()
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=9, color="666666")

        # Formate
        if "Nr." in label:
            val_cell.number_format = NUMBER_FORMAT_INT
        elif "%" in label and "RON" not in label:
            val_cell.number_format = NUMBER_FORMAT_PERCENT
        elif "R/R" in label:
            val_cell.number_format = '0.0'
        else:
            val_cell.number_format = NUMBER_FORMAT_CURRENCY

    # Alertă dacă poziția depășește 10%
    pct_row = current_row + 5
    ws.conditional_formatting.add(
        f"B{pct_row}",
        FormulaRule(formula=[f'B{pct_row}>{MAX_POSITION_PCT}'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # Alertă R/R sub minim
    rr_row = current_row + 6
    ws.conditional_formatting.add(
        f"B{rr_row}",
        FormulaRule(formula=[f'AND(B{rr_row}<>"",B{rr_row}<{MIN_RR_RATIO})'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    current_row += len(calc_results) + 2

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 3: ALOCARE SECTORIALĂ
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 3 — ALOCARE SECTORIALĂ", 5)
    current_row += 1

    sector_headers = ["Sector", "Valoare (RON)", "% Portofoliu", "Limită (30%)", "Status"]
    for c, h in enumerate(sector_headers, 1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center_align()
        cell.border = _border()
    current_row += 1

    sector_start = current_row
    for idx, sector in enumerate(SECTOARE[:10]):  # Top 10 sectoare
        r = current_row + idx
        ws.cell(row=r, column=1, value=sector).border = _border()

        # SUMIFS din Poziții pe sector
        val_cell = ws.cell(row=r, column=2)
        val_cell.value = (
            f"=SUMIFS('Poziții'!L3:L100,'Poziții'!E3:E100,A{r})"
        )
        val_cell.number_format = NUMBER_FORMAT_CURRENCY
        val_cell.border = _border()
        val_cell.alignment = _center_align()

        pct_cell = ws.cell(row=r, column=3)
        total_ref = r0  # row of total portfolio value from section 1
        pct_cell.value = f"=IF(B{r}>0,B{r}/SUM(B{sector_start}:B{sector_start+9}+Configurare!B4),0)"
        pct_cell.number_format = NUMBER_FORMAT_PERCENT
        pct_cell.border = _border()
        pct_cell.alignment = _center_align()

        ws.cell(row=r, column=4, value=MAX_SECTOR_PCT).border = _border()
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=4).alignment = _center_align()

        ws.cell(row=r, column=5).value = f'=IF(C{r}>{MAX_SECTOR_PCT},"DEPĂȘIT","OK")'
        ws.cell(row=r, column=5).border = _border()
        ws.cell(row=r, column=5).alignment = _center_align()
        ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True)

    sector_end = current_row + 9

    # Conditional formatting pe % sector
    ws.conditional_formatting.add(
        f"C{sector_start}:C{sector_end}",
        FormulaRule(formula=[f'AND(C{sector_start}<>"",C{sector_start}>{MAX_SECTOR_PCT})'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # Status DEPĂȘIT = roșu
    ws.conditional_formatting.add(
        f"E{sector_start}:E{sector_end}",
        FormulaRule(formula=[f'E{sector_start}="DEPĂȘIT"'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    current_row = sector_end + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 4: STOP-LOSS MONITORING
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 4 — MONITORIZARE STOP-LOSS", 6)
    current_row += 1

    sl_headers = ["Simbol", "Preț Curent", "Stop-Loss", "Distanță (RON)", "Distanță (%)", "Alertă"]
    for c, h in enumerate(sl_headers, 1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center_align()
        cell.border = _border()
    current_row += 1

    sl_start = current_row
    for idx in range(10):  # Max 10 poziții monitorizate
        r = current_row + idx
        data_row = 3 + idx

        ws.cell(row=r, column=1).value = f"=IF('Poziții'!B{data_row}<>\"\",'Poziții'!B{data_row},\"\")"
        ws.cell(row=r, column=2).value = f"=IF(A{r}<>\"\",'Poziții'!J{data_row},\"\")"
        ws.cell(row=r, column=2).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=3).value = f"=IF(A{r}<>\"\",'Poziții'!O{data_row},\"\")"
        ws.cell(row=r, column=3).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=4).value = f"=IF(A{r}<>\"\",B{r}-C{r},\"\")"
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=5).value = f"=IF(AND(A{r}<>\"\",B{r}>0),(B{r}-C{r})/B{r},\"\")"
        ws.cell(row=r, column=5).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=6).value = (
            f'=IF(A{r}="","",IF(E{r}<0.02,"CRITIC",IF(E{r}<0.05,"ATENȚIE","OK")))'
        )

        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _border()
            ws.cell(row=r, column=c).alignment = _center_align()

    sl_end = current_row + 9

    # Conditional formatting alertă SL
    ws.conditional_formatting.add(
        f"F{sl_start}:F{sl_end}",
        FormulaRule(formula=[f'F{sl_start}="CRITIC"'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )
    ws.conditional_formatting.add(
        f"F{sl_start}:F{sl_end}",
        FormulaRule(formula=[f'F{sl_start}="ATENȚIE"'],
                    fill=PatternFill(start_color=COLORS["light_yellow"],
                                    end_color=COLORS["light_yellow"],
                                    fill_type="solid"),
                    font=Font(color="FF8C00", bold=True))
    )

    current_row = sl_end + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 5: REFERINȚĂ REGULI RISK
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 5 — REGULI RISK MANAGEMENT (REFERINȚĂ)", 3)
    current_row += 1

    rules_text = [
        "1. Nicio poziție individuală nu depășește 10% din portofoliu (5% pt speculativ).",
        "2. Totalul pozițiilor pe un sector — maximum 30%.",
        "3. Cash reserve permanent: minimum 10-15%.",
        "4. Stop-loss obligatoriu pe fiecare poziție: -7% până la -10%.",
        "5. Trailing stop: breakeven la +15%, trailing -10% la +25%.",
        "6. R/R minim 2:1 (1.5:1 doar pe setup A+).",
        "7. Drawdown -10% → reduce expunerea la 50%.",
        "8. Drawdown -15% → ieșire pe cash 70-80%.",
        "9. După 3 SL consecutive → pauză 48h, re-analiză completă.",
        "10. NICIODATĂ nu muți stop-ul în jos. NICIODATĂ.",
        "11. NICIODATĂ averaging down fără teză fundamentală nouă.",
    ]

    for idx, rule in enumerate(rules_text):
        r = current_row + idx
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws.cell(row=r, column=1, value=rule)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        # Regulile critice (10, 11) în roșu bold
        if "NICIODATĂ" in rule:
            cell.font = Font(name="Calibri", size=10, bold=True,
                             color=COLORS["loss"])

    # Lățimi coloane
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    ws.freeze_panes = "A2"
    return ws


# ── Helper functions ──

def _section_header(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(num_cols, 1))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True,
                     color=COLORS["header_font"])
    cell.fill = PatternFill(start_color=COLORS["danger"],
                            end_color=COLORS["danger"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _header_font():
    return Font(name="Calibri", size=10, bold=True, color=COLORS["header_font"])


def _header_fill():
    return PatternFill(start_color=COLORS["header_bg"],
                       end_color=COLORS["header_bg"], fill_type="solid")


def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)
