"""
Foaia Analiză Tehnică — Workspace cu secțiuni: trend, momentum, volum, nivele, pattern-uri
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, ANALIZA_TEHNICA_DEMO, TREND, SEMNAL,
    NUMBER_FORMAT_PRICE, NUMBER_FORMAT_INT,
)


def create_sheet(wb):
    """Creează foaia Analiză Tehnică"""
    ws = wb.create_sheet(title="Analiză Tehnică")

    # Titlu principal
    ws.merge_cells("A1:N1")
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | ANALIZĂ TEHNICĂ — WORKSPACE")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 1: STRUCTURA TRENDULUI
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 1 — STRUCTURA TRENDULUI", 7)
    current_row += 1

    trend_headers = [
        ("Simbol", 10), ("Preț Curent", 14), ("MA50", 12), ("MA200", 12),
        ("Trend MA", 14), ("Golden/Death Cross", 18), ("Semnal", 16),
    ]
    _write_headers(ws, current_row, trend_headers)
    current_row += 1

    for i, at in enumerate(ANALIZA_TEHNICA_DEMO):
        r = current_row + i
        ws.cell(row=r, column=1, value=at["simbol"])
        ws.cell(row=r, column=2, value=at["pret"])
        ws.cell(row=r, column=3, value=at["ma50"])
        ws.cell(row=r, column=4, value=at["ma200"])
        ws.cell(row=r, column=5, value=at["trend_ma"])
        ws.cell(row=r, column=6, value=at["cross"])

        # Semnal automat bazat pe MA
        ws.cell(row=r, column=7).value = (
            f'=IF(AND(B{r}>C{r},B{r}>D{r}),"Bullish",'
            f'IF(AND(B{r}<C{r},B{r}<D{r}),"Bearish","Neutru"))'
        )

        _style_data_row(ws, r, len(trend_headers), i % 2 == 1)
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).number_format = NUMBER_FORMAT_PRICE

    # Conditional formatting — trend
    last_trend = current_row + len(ANALIZA_TEHNICA_DEMO) - 1
    _add_trend_formatting(ws, "E", current_row, last_trend)
    _add_trend_formatting(ws, "G", current_row, last_trend)

    # Golden Cross = verde, Death Cross = roșu
    ws.conditional_formatting.add(
        f"F{current_row}:F{last_trend}",
        CellIsRule(operator="equal", formula=['"Golden Cross"'],
                   fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                   font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"F{current_row}:F{last_trend}",
        CellIsRule(operator="equal", formula=['"Death Cross"'],
                   fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                   font=Font(color=COLORS["loss"], bold=True))
    )

    current_row = last_trend + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 2: INDICATORI MOMENTUM
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 2 — INDICATORI MOMENTUM", 9)
    current_row += 1

    mom_headers = [
        ("Simbol", 10), ("RSI (14)", 10), ("Semnal RSI", 14),
        ("MACD", 10), ("Semnal MACD", 14),
        ("Stoch %K", 10), ("Stoch %D", 10), ("Semnal Stoch", 14),
        ("Verdict Momentum", 18),
    ]
    _write_headers(ws, current_row, mom_headers)
    current_row += 1

    for i, at in enumerate(ANALIZA_TEHNICA_DEMO):
        r = current_row + i
        ws.cell(row=r, column=1, value=at["simbol"])
        ws.cell(row=r, column=2, value=at["rsi"])
        ws.cell(row=r, column=3, value=at["semnal_rsi"])
        ws.cell(row=r, column=4, value=at["macd"])
        ws.cell(row=r, column=5, value=at["semnal_macd"])
        ws.cell(row=r, column=6, value=at["stoch_k"])
        ws.cell(row=r, column=7, value=at["stoch_d"])
        ws.cell(row=r, column=8, value=at["semnal_stoch"])

        # Verdict momentum automat
        ws.cell(row=r, column=9).value = (
            f'=IF(AND(B{r}>50,D{r}>0),"Bullish",'
            f'IF(AND(B{r}<50,D{r}<0),"Bearish","Neutru"))'
        )

        _style_data_row(ws, r, len(mom_headers), i % 2 == 1)

    last_mom = current_row + len(ANALIZA_TEHNICA_DEMO) - 1

    # RSI conditional formatting: >70 = supracumpărat (roșu), <30 = supravândut (verde)
    ws.conditional_formatting.add(
        f"B{current_row}:B{last_mom}",
        FormulaRule(formula=[f'AND(B{current_row}<>"",B{current_row}>70)'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )
    ws.conditional_formatting.add(
        f"B{current_row}:B{last_mom}",
        FormulaRule(formula=[f'AND(B{current_row}<>"",B{current_row}<30)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"], bold=True))
    )

    _add_trend_formatting(ws, "I", current_row, last_mom)

    current_row = last_mom + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 3: ANALIZĂ VOLUM
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 3 — ANALIZĂ VOLUM", 6)
    current_row += 1

    vol_headers = [
        ("Simbol", 10), ("Volum Mediu (20z)", 16), ("Volum Curent", 14),
        ("Raport Volum", 14), ("OBV Trend", 14), ("Confirmare", 14),
    ]
    _write_headers(ws, current_row, vol_headers)
    current_row += 1

    for i, at in enumerate(ANALIZA_TEHNICA_DEMO):
        r = current_row + i
        ws.cell(row=r, column=1, value=at["simbol"])
        ws.cell(row=r, column=2, value=at["volum_mediu"])
        ws.cell(row=r, column=3, value=at["volum_curent"])
        ws.cell(row=r, column=4).value = f"=IF(B{r}>0,C{r}/B{r},0)"
        ws.cell(row=r, column=5, value=at["obv_trend"])
        ws.cell(row=r, column=6).value = (
            f'=IF(D{r}>1.5,"Volum Ridicat",IF(D{r}<0.7,"Volum Scăzut","Normal"))'
        )

        _style_data_row(ws, r, len(vol_headers), i % 2 == 1)
        ws.cell(row=r, column=2).number_format = NUMBER_FORMAT_INT
        ws.cell(row=r, column=3).number_format = NUMBER_FORMAT_INT
        ws.cell(row=r, column=4).number_format = '0.00x'

    last_vol = current_row + len(ANALIZA_TEHNICA_DEMO) - 1
    current_row = last_vol + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 4: NIVELURI CHEIE
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 4 — NIVELURI CHEIE (Suport/Rezistență + Fibonacci)", 10)
    current_row += 1

    level_headers = [
        ("Simbol", 10), ("Suport 1", 12), ("Suport 2", 12),
        ("Rezistență 1", 14), ("Rezistență 2", 14),
        ("Fib 23.6%", 12), ("Fib 38.2%", 12), ("Fib 50%", 10),
        ("Fib 61.8%", 12), ("Preț vs Niveluri", 18),
    ]
    _write_headers(ws, current_row, level_headers)
    current_row += 1

    for i, at in enumerate(ANALIZA_TEHNICA_DEMO):
        r = current_row + i
        ws.cell(row=r, column=1, value=at["simbol"])
        ws.cell(row=r, column=2, value=at["suport1"])
        ws.cell(row=r, column=3, value=at["suport2"])
        ws.cell(row=r, column=4, value=at["rezistenta1"])
        ws.cell(row=r, column=5, value=at["rezistenta2"])

        # Fibonacci auto: bazat pe Suport2 (low) și Rezistenta2 (high)
        ws.cell(row=r, column=6).value = f"=E{r}-(E{r}-C{r})*0.236"  # 23.6%
        ws.cell(row=r, column=7).value = f"=E{r}-(E{r}-C{r})*0.382"  # 38.2%
        ws.cell(row=r, column=8).value = f"=E{r}-(E{r}-C{r})*0.5"    # 50%
        ws.cell(row=r, column=9).value = f"=E{r}-(E{r}-C{r})*0.618"  # 61.8%

        # Preț vs niveluri (referință la secțiunea 1, coloana B, la rândul corespunzător)
        pret_ref = at["pret"]
        ws.cell(row=r, column=10).value = (
            f'=IF({pret_ref}>D{r},"Deasupra R1",'
            f'IF({pret_ref}<B{r},"Sub S1","Între S1-R1"))'
        )

        _style_data_row(ws, r, len(level_headers), i % 2 == 1)
        for c in range(2, 10):
            ws.cell(row=r, column=c).number_format = NUMBER_FORMAT_PRICE

    last_level = current_row + len(ANALIZA_TEHNICA_DEMO) - 1
    current_row = last_level + 3

    # ═══════════════════════════════════════════════════════
    # SECȚIUNEA 5: PATTERN DETECTION
    # ═══════════════════════════════════════════════════════
    _section_header(ws, current_row, "SECȚIUNEA 5 — PATTERN-URI DETECTATE", 6)
    current_row += 1

    pat_headers = [
        ("Simbol", 10), ("Pattern Detectat", 22), ("Timeframe", 14),
        ("Direcție", 12), ("Confirmat", 12), ("Observații", 40),
    ]
    _write_headers(ws, current_row, pat_headers)
    current_row += 1

    for i, at in enumerate(ANALIZA_TEHNICA_DEMO):
        r = current_row + i
        ws.cell(row=r, column=1, value=at["simbol"])
        ws.cell(row=r, column=2, value=at["pattern"])
        ws.cell(row=r, column=3, value=at["pattern_tf"])
        ws.cell(row=r, column=4, value=at["pattern_dir"])
        ws.cell(row=r, column=5, value="Da" if at["pattern_dir"] != "Neutru" else "Parțial")
        ws.cell(row=r, column=6, value="")

        _style_data_row(ws, r, len(pat_headers), i % 2 == 1)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="left",
                                                        vertical="center",
                                                        wrap_text=True)

    _add_trend_formatting(ws, "D", current_row,
                          current_row + len(ANALIZA_TEHNICA_DEMO) - 1)

    # Lățimi coloane principale
    for col in range(1, 15):
        if ws.column_dimensions[get_column_letter(col)].width == 0:
            ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "A2"
    return ws


# ── Helper functions ──

def _section_header(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(num_cols, 1))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True,
                     color=COLORS["header_font"])
    cell.fill = PatternFill(start_color=COLORS["purple"],
                            end_color=COLORS["purple"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_headers(ws, row, headers):
    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = Font(name="Calibri", size=10, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            width, ws.column_dimensions[get_column_letter(col_idx)].width or 0
        )


def _style_data_row(ws, row, num_cols, is_alt):
    fill = PatternFill(start_color=COLORS["row_alt"],
                       end_color=COLORS["row_alt"],
                       fill_type="solid") if is_alt else None
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = _border()
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _add_trend_formatting(ws, col, start_row, end_row):
    """Bullish = verde, Bearish = roșu pe o coloană"""
    cell_range = f"{col}{start_row}:{col}{end_row}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Bullish"'],
                   fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                   font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"Bearish"'],
                   fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                   font=Font(color=COLORS["loss"], bold=True))
    )
