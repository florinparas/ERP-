"""
Foaia Poziții — Toate pozițiile deschise cu P&L live
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config.stock_config import (
    COLORS, POZITII_DEMO, SECTOARE, TIP_POZITIE, STATUS_POZITIE,
    PIATA, MONEDA, NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT,
    NUMBER_FORMAT_PRICE, NUMBER_FORMAT_INT,
)


def create_sheet(wb):
    """Creează foaia Poziții"""
    ws = wb.create_sheet(title="Poziții")

    headers = [
        ("ID", 8),
        ("Simbol", 10),
        ("Denumire", 28),
        ("Piață", 10),
        ("Sector", 18),
        ("Tip", 14),
        ("Cantitate", 12),
        ("Preț Intrare", 14),
        ("Data Intrare", 14),
        ("Preț Curent", 14),
        ("Val. Investită", 16),
        ("Val. Curentă", 16),
        ("P&L (RON)", 14),
        ("P&L (%)", 11),
        ("Stop-Loss", 12),
        ("Dist. SL (%)", 12),
        ("Target 1", 12),
        ("Target 2", 12),
        ("R/R Ratio", 11),
        ("Status", 14),
        ("Monedă", 10),
        ("Observații", 35),
    ]

    num_cols = len(headers)

    # Titlu
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | POZIȚII DESCHISE")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headere
    for col_idx, (col_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=11, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Date demo
    for row_offset, pos in enumerate(POZITII_DEMO):
        r = 3 + row_offset
        ws.cell(row=r, column=1, value=pos["id"])
        ws.cell(row=r, column=2, value=pos["simbol"])
        ws.cell(row=r, column=3, value=pos["denumire"])
        ws.cell(row=r, column=4, value=pos["piata"])
        ws.cell(row=r, column=5, value=pos["sector"])
        ws.cell(row=r, column=6, value=pos["tip"])
        ws.cell(row=r, column=7, value=pos["cantitate"])
        ws.cell(row=r, column=8, value=pos["pret_intrare"])
        ws.cell(row=r, column=9, value=pos["data_intrare"])
        ws.cell(row=r, column=10, value=pos["pret_curent"])

        # Formule
        ws.cell(row=r, column=11).value = f"=G{r}*H{r}"        # Val. Investită
        ws.cell(row=r, column=12).value = f"=G{r}*J{r}"        # Val. Curentă
        ws.cell(row=r, column=13).value = f"=L{r}-K{r}"        # P&L RON
        ws.cell(row=r, column=14).value = f"=IF(K{r}>0,(L{r}-K{r})/K{r},0)"  # P&L %
        ws.cell(row=r, column=15, value=pos["stop_loss"])
        ws.cell(row=r, column=16).value = f"=IF(J{r}>0,(J{r}-O{r})/J{r},0)"  # Dist. SL %
        ws.cell(row=r, column=17, value=pos["target1"])
        ws.cell(row=r, column=18, value=pos["target2"])
        ws.cell(row=r, column=19).value = (
            f"=IF(H{r}-O{r}>0,(Q{r}-H{r})/(H{r}-O{r}),0)"
        )  # R/R Ratio
        ws.cell(row=r, column=20, value="Deschisă")
        ws.cell(row=r, column=21, value=pos["moneda"])

        # Formatare rânduri
        is_alt = row_offset % 2 == 1
        fill = PatternFill(start_color=COLORS["row_alt"],
                           end_color=COLORS["row_alt"],
                           fill_type="solid") if is_alt else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10)
            if fill:
                cell.fill = fill

        # Formate numerice specifice
        ws.cell(row=r, column=7).number_format = NUMBER_FORMAT_INT
        ws.cell(row=r, column=8).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=10).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=11).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=12).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=13).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=14).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=15).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=16).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=17).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=18).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=19).number_format = '0.0'

    # Rând TOTAL
    total_row = 3 + len(POZITII_DEMO)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", size=11, bold=True)

    ws.cell(row=total_row, column=11).value = f"=SUM(K3:K{total_row - 1})"
    ws.cell(row=total_row, column=11).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=total_row, column=12).value = f"=SUM(L3:L{total_row - 1})"
    ws.cell(row=total_row, column=12).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=total_row, column=13).value = f"=SUM(M3:M{total_row - 1})"
    ws.cell(row=total_row, column=13).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=total_row, column=14).value = (
        f"=IF(K{total_row}>0,M{total_row}/K{total_row},0)"
    )
    ws.cell(row=total_row, column=14).number_format = NUMBER_FORMAT_PERCENT

    for c in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.border = _border()
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color=COLORS["light_blue"],
                                end_color=COLORS["light_blue"], fill_type="solid")

    # Conditional formatting — P&L verde/roșu
    green_fill = PatternFill(start_color=COLORS["light_green"],
                             end_color=COLORS["light_green"], fill_type="solid")
    red_fill = PatternFill(start_color=COLORS["light_red"],
                           end_color=COLORS["light_red"], fill_type="solid")

    # P&L RON (col M)
    ws.conditional_formatting.add(
        f"M3:M{total_row - 1}",
        FormulaRule(formula=[f'AND(M3<>"",M3>0)'], fill=green_fill,
                    font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"M3:M{total_row - 1}",
        FormulaRule(formula=[f'AND(M3<>"",M3<0)'], fill=red_fill,
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # P&L % (col N)
    ws.conditional_formatting.add(
        f"N3:N{total_row - 1}",
        FormulaRule(formula=[f'AND(N3<>"",N3>0)'], fill=green_fill,
                    font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"N3:N{total_row - 1}",
        FormulaRule(formula=[f'AND(N3<>"",N3<0)'], fill=red_fill,
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # Alertă stop-loss — galben dacă distanța < 5%, roșu dacă < 2%
    yellow_fill = PatternFill(start_color=COLORS["light_yellow"],
                              end_color=COLORS["light_yellow"], fill_type="solid")
    ws.conditional_formatting.add(
        f"P3:P{total_row - 1}",
        FormulaRule(formula=[f'AND(P3<>"",P3<0.02)'], fill=red_fill,
                    font=Font(color=COLORS["loss"], bold=True))
    )
    ws.conditional_formatting.add(
        f"P3:P{total_row - 1}",
        FormulaRule(formula=[f'AND(P3<>"",P3<0.05,P3>=0.02)'], fill=yellow_fill,
                    font=Font(color="FF8C00", bold=True))
    )

    # Validări dropdown
    _add_dropdown(ws, f"D3:D1000", PIATA)
    _add_dropdown(ws, f"E3:E1000", SECTOARE)
    _add_dropdown(ws, f"F3:F1000", TIP_POZITIE)
    _add_dropdown(ws, f"T3:T1000", STATUS_POZITIE)
    _add_dropdown(ws, f"U3:U1000", MONEDA)

    # Freeze panes & filter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

    return ws


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _add_dropdown(ws, cell_range, options):
    formula = '"' + ','.join(str(o) for o in options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Selectați din lista disponibilă."
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
