"""
Foaia Configurare — Parametri portofoliu, reguli risk, liste dropdown, comisioane
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, PORTFOLIO_INITIAL_CAPITAL, PORTFOLIO_CASH, PORTFOLIO_START_DATE,
    MAX_POSITION_PCT, MAX_POSITION_SPECULATIVE_PCT, MAX_SECTOR_PCT,
    MIN_CASH_RESERVE_PCT, STOP_LOSS_DEFAULT_PCT, STOP_LOSS_MAX_PCT,
    TRAILING_STOP_BREAKEVEN_TRIGGER, TRAILING_STOP_TRAIL_TRIGGER,
    TRAILING_STOP_TRAIL_PCT, MIN_RR_RATIO, MIN_RR_SPECULATIVE,
    DRAWDOWN_REDUCE_PCT, DRAWDOWN_EXIT_PCT, MAX_CONSECUTIVE_LOSSES,
    SECTOARE, TIP_TRANZACTIE, TIP_POZITIE, STATUS_POZITIE, CONVINGERE,
    TIMEFRAME, PIATA, MONEDA, TREND, SEMNAL, VERDICT_FUNDAMENTAL,
    COMISIOANE, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_CURRENCY,
)


def create_sheet(wb):
    """Creează foaia Configurare"""
    ws = wb.create_sheet(title="Configurare")

    # ---- SECȚIUNEA 1: PARAMETRI PORTOFOLIU ----
    _section_header(ws, 1, "PARAMETRI PORTOFOLIU", 3)

    params = [
        ("Parametru", "Valoare", "Descriere"),
        ("Capital Inițial", PORTFOLIO_INITIAL_CAPITAL, "Capital total investit inițial (RON)"),
        ("Cash Disponibil", PORTFOLIO_CASH, "Cash curent disponibil (RON)"),
        ("Data Start", PORTFOLIO_START_DATE, "Data deschidere portofoliu"),
        ("Moneda Bază", "RON", "Moneda de referință a portofoliului"),
    ]

    for row_idx, (param, val, desc) in enumerate(params, 2):
        ws.cell(row=row_idx, column=1, value=param)
        cell_val = ws.cell(row=row_idx, column=2, value=val)
        ws.cell(row=row_idx, column=3, value=desc)

        if row_idx == 2:
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = _header_font()
                cell.fill = _header_fill()
                cell.alignment = _center_align()
                cell.border = _border()
        else:
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = _border()
                cell.alignment = _left_align()
            if isinstance(val, (int, float)) and val >= 100:
                cell_val.number_format = NUMBER_FORMAT_CURRENCY

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45

    # ---- SECȚIUNEA 2: REGULI RISK MANAGEMENT ----
    risk_start = 8
    _section_header(ws, risk_start, "REGULI RISK MANAGEMENT — FĂRĂ EXCEPȚII", 3)

    risk_rules = [
        ("Regulă", "Valoare", "Descriere"),
        ("Max Poziție Standard", MAX_POSITION_PCT, "Maximum 10% din portofoliu per poziție"),
        ("Max Poziție Speculativă", MAX_POSITION_SPECULATIVE_PCT, "Maximum 5% din portofoliu per poziție speculativă"),
        ("Max Sector", MAX_SECTOR_PCT, "Maximum 30% din portofoliu per sector"),
        ("Cash Minim", MIN_CASH_RESERVE_PCT, "Rezervă cash permanentă minimum 10%"),
        ("Stop-Loss Standard", STOP_LOSS_DEFAULT_PCT, "Stop-loss default: 7% sub preț intrare"),
        ("Stop-Loss Maxim", STOP_LOSS_MAX_PCT, "Stop-loss maxim acceptat: 10%"),
        ("Trailing SL → Breakeven", TRAILING_STOP_BREAKEVEN_TRIGGER, "Mută SL la breakeven după +15% profit"),
        ("Trailing SL → Trail", TRAILING_STOP_TRAIL_TRIGGER, "Activează trailing -10% după +25% profit"),
        ("Trailing Stop %", TRAILING_STOP_TRAIL_PCT, "Trailing stop: 10% sub maximul atins"),
        ("R/R Minim Standard", MIN_RR_RATIO, "Risk/Reward minim: 2:1"),
        ("R/R Minim Speculativ", MIN_RR_SPECULATIVE, "R/R minim speculativ (doar setup A+): 1.5:1"),
        ("Drawdown → Reduce", DRAWDOWN_REDUCE_PCT, "Reduce expunerea la 50% dacă portofoliul scade -10%"),
        ("Drawdown → Exit", DRAWDOWN_EXIT_PCT, "Ieșire pe cash 70-80% dacă portofoliul scade -15%"),
        ("Max SL Consecutive", MAX_CONSECUTIVE_LOSSES, "Pauză 48h după 3 stop-loss-uri consecutive"),
    ]

    for row_idx, (param, val, desc) in enumerate(risk_rules, risk_start + 1):
        ws.cell(row=row_idx, column=1, value=param)
        cell_val = ws.cell(row=row_idx, column=2, value=val)
        ws.cell(row=row_idx, column=3, value=desc)

        if row_idx == risk_start + 1:
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = _header_font()
                cell.fill = _header_fill()
                cell.alignment = _center_align()
                cell.border = _border()
        else:
            for c in range(1, 4):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = _border()
                cell.alignment = _left_align()
            if isinstance(val, float) and val < 1:
                cell_val.number_format = NUMBER_FORMAT_PERCENT
            elif isinstance(val, float) and val >= 1:
                cell_val.number_format = '0.0":1"'

    # ---- SECȚIUNEA 3: COMISIOANE BROKERI ----
    com_start = risk_start + len(risk_rules) + 2
    _section_header(ws, com_start, "COMISIOANE BROKERI", 4)

    com_headers = ["Piață", "Tip", "Valoare", "Minim"]
    for c, h in enumerate(com_headers, 1):
        cell = ws.cell(row=com_start + 1, column=c, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center_align()
        cell.border = _border()

    for row_idx, (piata, info) in enumerate(COMISIOANE.items(), com_start + 2):
        ws.cell(row=row_idx, column=1, value=piata).border = _border()
        ws.cell(row=row_idx, column=2, value=info["tip"]).border = _border()
        val_cell = ws.cell(row=row_idx, column=3, value=info["valoare"])
        val_cell.border = _border()
        if info["tip"] == "procent":
            val_cell.number_format = NUMBER_FORMAT_PERCENT
        min_cell = ws.cell(row=row_idx, column=4,
                           value=f'{info["minim"]:.2f} {info["moneda"]}')
        min_cell.border = _border()

    ws.column_dimensions["D"].width = 15

    # ---- SECȚIUNEA 4: LISTE DROPDOWN (coloanele F+) ----
    lists_col_start = 6
    dropdown_lists = [
        ("Sectoare", SECTOARE),
        ("Tip Tranzacție", TIP_TRANZACTIE),
        ("Tip Poziție", TIP_POZITIE),
        ("Status Poziție", STATUS_POZITIE),
        ("Convingere", CONVINGERE),
        ("Timeframe", TIMEFRAME),
        ("Piață", PIATA),
        ("Monedă", MONEDA),
        ("Trend", TREND),
        ("Semnal", SEMNAL),
        ("Verdict", VERDICT_FUNDAMENTAL),
    ]

    _section_header(ws, 1, "LISTE DROPDOWN — VALIDĂRI",
                    len(dropdown_lists), start_col=lists_col_start)

    for col_offset, (list_name, items) in enumerate(dropdown_lists):
        col = lists_col_start + col_offset
        header_cell = ws.cell(row=2, column=col, value=list_name)
        header_cell.font = _header_font()
        header_cell.fill = _header_fill()
        header_cell.alignment = _center_align()
        header_cell.border = _border()

        for row_idx, item in enumerate(items, 3):
            cell = ws.cell(row=row_idx, column=col, value=item)
            cell.border = _border()
            cell.alignment = _left_align()

        ws.column_dimensions[get_column_letter(col)].width = max(
            len(list_name) + 2,
            max(len(str(i)) for i in items) + 2
        )

    ws.freeze_panes = "A3"
    return ws


def _section_header(ws, row, title, num_cols, start_col=1):
    """Adaugă un header de secțiune"""
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row, end_column=start_col + num_cols - 1
    )
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["header_font"])
    cell.fill = PatternFill(
        start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _header_font():
    return Font(name="Calibri", size=11, bold=True, color=COLORS["header_font"])


def _header_fill():
    return PatternFill(start_color=COLORS["header_bg"],
                       end_color=COLORS["header_bg"], fill_type="solid")


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)
