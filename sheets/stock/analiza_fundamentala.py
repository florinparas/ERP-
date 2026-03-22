"""
Foaia Analiză Fundamentală — Screening fundamental companii
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config.stock_config import (
    COLORS, ANALIZA_FUNDAMENTALA_DEMO, SECTOARE, VERDICT_FUNDAMENTAL,
    NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_PRICE,
    NUMBER_FORMAT_INT,
)


def create_sheet(wb):
    """Creează foaia Analiză Fundamentală"""
    ws = wb.create_sheet(title="Analiză Fundamentală")

    headers = [
        ("Simbol", 10),
        ("Denumire", 25),
        ("Sector", 16),
        ("Preț", 12),
        ("Capitaliz. (mil)", 14),
        ("Revenue YoY%", 14),
        ("EPS Growth%", 14),
        ("P/E", 10),
        ("P/E Sector", 12),
        ("EV/EBITDA", 12),
        ("FCF Yield%", 12),
        ("ROE%", 10),
        ("Debt/Equity", 12),
        ("Div. Yield%", 12),
        ("Payout%", 10),
        ("Scor (1-10)", 12),
        ("Verdict", 14),
        ("Observații", 40),
    ]

    num_cols = len(headers)

    # Titlu
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | ANALIZĂ FUNDAMENTALĂ — SCREENING")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headere
    for col_idx, (col_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Date demo
    observatii_map = {
        "TLV": "Blue chip solid. P/E sub media sectorului, ROE excelent. BUY pe orice pullback.",
        "SNG": "Cash cow. Dividend yield ridicat, debt scăzut. Poziție defensivă de bază.",
        "SNP": "EPS în scădere — atenție. Dependent de prețul petrolului. HOLD cu SL strâns.",
        "FP": "Discount NAV masiv. Cel mai ieftin vehicle de investiții pe BVB. Strong BUY.",
        "BRD": "Solid dar fără catalizator imediat. P/E corect. HOLD.",
        "DIGI": "Growth story puternică. Leverage ridicat dar justificat de CAPEX. Risc/recompensă OK.",
    }

    for row_offset, af in enumerate(ANALIZA_FUNDAMENTALA_DEMO):
        r = 3 + row_offset
        ws.cell(row=r, column=1, value=af["simbol"])
        ws.cell(row=r, column=2, value=af["denumire"])
        ws.cell(row=r, column=3, value=af["sector"])
        ws.cell(row=r, column=4, value=af["pret"])
        ws.cell(row=r, column=5, value=af["capitalizare"])
        ws.cell(row=r, column=6, value=af["revenue_growth"] / 100)
        ws.cell(row=r, column=7, value=af["eps_growth"] / 100)
        ws.cell(row=r, column=8, value=af["pe"])
        ws.cell(row=r, column=9, value=af["pe_sector"])
        ws.cell(row=r, column=10, value=af["ev_ebitda"])
        ws.cell(row=r, column=11, value=af["fcf_yield"] / 100)
        ws.cell(row=r, column=12, value=af["roe"] / 100)
        ws.cell(row=r, column=13, value=af["debt_equity"])
        ws.cell(row=r, column=14, value=af["div_yield"] / 100)
        ws.cell(row=r, column=15, value=af["payout"] / 100)
        ws.cell(row=r, column=16, value=af["scor"])

        # Verdict automat
        ws.cell(row=r, column=17).value = (
            f'=IF(P{r}>=8,"BUY",IF(P{r}>=6,"HOLD",IF(P{r}>=4,"WATCH","SELL")))'
        )

        ws.cell(row=r, column=18, value=observatii_map.get(af["simbol"], ""))

        # Stilizare
        is_alt = row_offset % 2 == 1
        fill = PatternFill(start_color=COLORS["row_alt"],
                           end_color=COLORS["row_alt"],
                           fill_type="solid") if is_alt else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _border()
            cell.font = Font(name="Calibri", size=10)
            if c == 18:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

        # Formate numerice
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=5).number_format = NUMBER_FORMAT_INT
        ws.cell(row=r, column=6).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=7).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=8).number_format = '0.0'
        ws.cell(row=r, column=9).number_format = '0.0'
        ws.cell(row=r, column=10).number_format = '0.0'
        ws.cell(row=r, column=11).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=12).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=13).number_format = '0.00'
        ws.cell(row=r, column=14).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=15).number_format = NUMBER_FORMAT_PERCENT

    last_data_row = 2 + len(ANALIZA_FUNDAMENTALA_DEMO)

    # ── Conditional Formatting ──

    # P/E vs Sector: verde dacă sub media sectorului, roșu dacă peste
    ws.conditional_formatting.add(
        f"H3:H{last_data_row}",
        FormulaRule(formula=[f'AND(H3<>"",H3<I3)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"H3:H{last_data_row}",
        FormulaRule(formula=[f'AND(H3<>"",H3>I3)'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"]))
    )

    # ROE: verde dacă > 15%
    ws.conditional_formatting.add(
        f"L3:L{last_data_row}",
        FormulaRule(formula=[f'AND(L3<>"",L3>0.15)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"], bold=True))
    )

    # Debt/Equity: roșu dacă > 1.5, galben dacă > 1.0, verde dacă < 0.5
    ws.conditional_formatting.add(
        f"M3:M{last_data_row}",
        FormulaRule(formula=[f'AND(M3<>"",M3>1.5)'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )
    ws.conditional_formatting.add(
        f"M3:M{last_data_row}",
        FormulaRule(formula=[f'AND(M3<>"",M3<=0.5)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"]))
    )

    # Scor fundamental gradient (1-10)
    score_colors = {
        (1, 3): (COLORS["light_red"], COLORS["loss"]),
        (4, 5): (COLORS["light_yellow"], "FF8C00"),
        (6, 7): (COLORS["light_blue"], COLORS["accent"]),
        (8, 10): (COLORS["light_green"], COLORS["profit"]),
    }
    for (low, high), (bg, fg) in score_colors.items():
        for score in range(low, high + 1):
            ws.conditional_formatting.add(
                f"P3:P{last_data_row}",
                CellIsRule(operator="equal", formula=[str(score)],
                           fill=PatternFill(start_color=bg, end_color=bg,
                                            fill_type="solid"),
                           font=Font(color=fg, bold=True))
            )

    # Verdict: BUY=verde, SELL=roșu, HOLD=albastru, WATCH=galben
    verdict_colors = {
        "BUY": (COLORS["light_green"], COLORS["profit"]),
        "SELL": (COLORS["light_red"], COLORS["loss"]),
        "HOLD": (COLORS["light_blue"], COLORS["accent"]),
        "WATCH": (COLORS["light_yellow"], "FF8C00"),
    }
    for verdict, (bg, fg) in verdict_colors.items():
        ws.conditional_formatting.add(
            f"Q3:Q{last_data_row}",
            CellIsRule(operator="equal", formula=[f'"{verdict}"'],
                       fill=PatternFill(start_color=bg, end_color=bg,
                                        fill_type="solid"),
                       font=Font(color=fg, bold=True))
        )

    # Validări
    _add_dropdown(ws, "C3:C1000", SECTOARE)
    _add_dropdown(ws, "Q3:Q1000", VERDICT_FUNDAMENTAL)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

    return ws


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _add_dropdown(ws, cell_range, options):
    formula = '"' + ','.join(str(o) for o in options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
