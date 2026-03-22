"""
Foaia Watchlist — Acțiuni monitorizate cu rating convingere
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, WATCHLIST_DEMO, SECTOARE, PIATA, TIMEFRAME,
    CONVINGERE_NUMERIC, NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT,
    NUMBER_FORMAT_PRICE,
)


def create_sheet(wb):
    """Creează foaia Watchlist"""
    ws = wb.create_sheet(title="Watchlist")

    headers = [
        ("Simbol", 10),
        ("Denumire", 25),
        ("Sector", 18),
        ("Piață", 10),
        ("Preț Curent", 14),
        ("Preț Țintă Intrare", 16),
        ("Distanță (%)", 12),
        ("SL Planificat", 14),
        ("Target", 12),
        ("R/R Potențial", 12),
        ("Convingere", 12),
        ("Timeframe", 20),
        ("Catalizator", 35),
        ("Data Adăugare", 14),
        ("Observații StockAgent", 50),
    ]

    num_cols = len(headers)

    # Titlu
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | WATCHLIST — OPORTUNITĂȚI MONITORIZATE")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headere
    for col_idx, (col_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=11, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Date demo
    for row_offset, w in enumerate(WATCHLIST_DEMO):
        r = 3 + row_offset
        ws.cell(row=r, column=1, value=w["simbol"])
        ws.cell(row=r, column=2, value=w["denumire"])
        ws.cell(row=r, column=3, value=w["sector"])
        ws.cell(row=r, column=4, value=w["piata"])
        ws.cell(row=r, column=5, value=w["pret_curent"])
        ws.cell(row=r, column=6, value=w["pret_tinta"])
        # Distanța % (preț curent vs preț țintă)
        ws.cell(row=r, column=7).value = f"=IF(E{r}>0,(E{r}-F{r})/E{r},0)"
        ws.cell(row=r, column=8, value=w["stop_loss"])
        ws.cell(row=r, column=9, value=w["target"])
        # R/R potențial
        ws.cell(row=r, column=10).value = (
            f"=IF(F{r}-H{r}>0,(I{r}-F{r})/(F{r}-H{r}),0)"
        )
        ws.cell(row=r, column=11, value=w["convingere"])
        ws.cell(row=r, column=12, value=w["timeframe"])
        ws.cell(row=r, column=13, value=w["catalizator"])
        ws.cell(row=r, column=14, value="2025-03-20")
        ws.cell(row=r, column=15, value=w["observatii"])

        # Stilizare
        is_alt = row_offset % 2 == 1
        fill = PatternFill(start_color=COLORS["row_alt"],
                           end_color=COLORS["row_alt"],
                           fill_type="solid") if is_alt else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _border()
            cell.font = Font(name="Calibri", size=10)
            if c in (13, 15):
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

        # Formate numerice
        ws.cell(row=r, column=5).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=6).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=7).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=8).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=9).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=10).number_format = '0.0'

    last_data_row = 2 + len(WATCHLIST_DEMO)

    # Conditional formatting — Convingere (gradient 1-5)
    score_colors = {
        1: ("FF0000", "FFFFFF"),
        2: ("FF8C00", "000000"),
        3: ("FFD700", "000000"),
        4: ("90EE90", "000000"),
        5: ("008000", "FFFFFF"),
    }
    for score, (bg_color, font_color) in score_colors.items():
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        rule = CellIsRule(
            operator="equal",
            formula=[str(score)],
            fill=fill,
            font=Font(name="Calibri", size=10, bold=True, color=font_color)
        )
        ws.conditional_formatting.add(f"K3:K{last_data_row}", rule)

    # R/R potențial — verde dacă > 3, galben dacă 2-3, roșu dacă < 2
    ws.conditional_formatting.add(
        f"J3:J{last_data_row}",
        FormulaRule(formula=[f'AND(J3<>"",J3>=3)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"J3:J{last_data_row}",
        FormulaRule(formula=[f'AND(J3<>"",J3<2)'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # Validări dropdown
    _add_dropdown(ws, "C3:C1000", SECTOARE)
    _add_dropdown(ws, "D3:D1000", PIATA)
    _add_dropdown(ws, "K3:K1000", CONVINGERE_NUMERIC)
    _add_dropdown(ws, "L3:L1000", TIMEFRAME)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

    return ws


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _add_dropdown(ws, cell_range, options):
    formula = '"' + ','.join(str(o) for o in options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
