"""
Foaia Tranzacții — Jurnal complet al tuturor tranzacțiilor
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, TRANZACTII_DEMO, TIP_TRANZACTIE,
    NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_PRICE,
    NUMBER_FORMAT_INT,
)


def create_sheet(wb):
    """Creează foaia Tranzacții"""
    ws = wb.create_sheet(title="Tranzacții")

    headers = [
        ("ID", 8),
        ("Data Intrare", 14),
        ("Simbol", 10),
        ("Denumire", 25),
        ("Tip", 14),
        ("Cantitate", 12),
        ("Preț Intrare", 14),
        ("Data Ieșire", 14),
        ("Preț Ieșire", 14),
        ("Val. Intrare", 15),
        ("Val. Ieșire", 15),
        ("P&L (RON)", 14),
        ("P&L (%)", 11),
        ("Comision", 12),
        ("P&L Net", 14),
        ("R/R Realizat", 12),
        ("Durată (zile)", 12),
        ("Motiv Intrare", 30),
        ("Motiv Ieșire", 30),
        ("Lecții Învățate", 40),
    ]

    num_cols = len(headers)

    # Titlu
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | JURNAL TRANZACȚII")
    title_cell.font = Font(name="Calibri", size=14, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headere
    for col_idx, (col_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=11, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Date demo
    for row_offset, tx in enumerate(TRANZACTII_DEMO):
        r = 3 + row_offset
        ws.cell(row=r, column=1, value=tx["id"])
        ws.cell(row=r, column=2, value=tx["data_intrare"])
        ws.cell(row=r, column=3, value=tx["simbol"])
        ws.cell(row=r, column=4, value=tx["denumire"])
        ws.cell(row=r, column=5, value=tx["tip"])
        ws.cell(row=r, column=6, value=tx["cantitate"])
        ws.cell(row=r, column=7, value=tx["pret_intrare"])
        ws.cell(row=r, column=8, value=tx["data_iesire"])
        ws.cell(row=r, column=9, value=tx["pret_iesire"])

        # Formule calculate
        ws.cell(row=r, column=10).value = f"=F{r}*G{r}"             # Val. Intrare
        ws.cell(row=r, column=11).value = f"=F{r}*I{r}"             # Val. Ieșire
        ws.cell(row=r, column=12).value = f"=K{r}-J{r}"             # P&L RON
        ws.cell(row=r, column=13).value = f"=IF(J{r}>0,(K{r}-J{r})/J{r},0)"  # P&L %
        ws.cell(row=r, column=14, value=tx["comision"])
        ws.cell(row=r, column=15).value = f"=L{r}-N{r}"             # P&L Net
        ws.cell(row=r, column=16).value = (
            f"=IF(AND(G{r}>0,I{r}>0),"
            f"ABS(I{r}-G{r})/ABS(G{r}-0),0)"
        )
        # R/R simplificat: |profit per share| / |risk per share|
        # Se calculează ca |pret_iesire - pret_intrare| / riscul estimat
        ws.cell(row=r, column=16).value = (
            f"=IF(J{r}>0,ABS(L{r})/ABS(J{r}*0.08),0)"
        )
        ws.cell(row=r, column=17).value = (
            f'=IF(AND(B{r}<>"",H{r}<>""),DATEVALUE(H{r})-DATEVALUE(B{r}),"")'
        )
        ws.cell(row=r, column=18, value=tx["motiv_intrare"])
        ws.cell(row=r, column=19, value=tx["motiv_iesire"])
        ws.cell(row=r, column=20, value=tx["lectii"])

        # Formatare
        is_alt = row_offset % 2 == 1
        fill = PatternFill(start_color=COLORS["row_alt"],
                           end_color=COLORS["row_alt"],
                           fill_type="solid") if is_alt else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _border()
            cell.font = Font(name="Calibri", size=10)
            if c <= 17:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)
            if fill:
                cell.fill = fill

        # Formate numerice
        ws.cell(row=r, column=6).number_format = NUMBER_FORMAT_INT
        ws.cell(row=r, column=7).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=9).number_format = NUMBER_FORMAT_PRICE
        ws.cell(row=r, column=10).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=11).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=12).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=13).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=14).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=15).number_format = NUMBER_FORMAT_CURRENCY
        ws.cell(row=r, column=16).number_format = '0.0'

    last_data_row = 2 + len(TRANZACTII_DEMO)

    # P&L Conditional formatting
    green_fill = PatternFill(start_color=COLORS["light_green"],
                             end_color=COLORS["light_green"], fill_type="solid")
    red_fill = PatternFill(start_color=COLORS["light_red"],
                           end_color=COLORS["light_red"], fill_type="solid")

    for col_letter in ["L", "M", "O"]:
        ws.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{last_data_row}",
            FormulaRule(formula=[f'AND({col_letter}3<>"",{col_letter}3>0)'],
                        fill=green_fill,
                        font=Font(color=COLORS["profit"], bold=True))
        )
        ws.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{last_data_row}",
            FormulaRule(formula=[f'AND({col_letter}3<>"",{col_letter}3<0)'],
                        fill=red_fill,
                        font=Font(color=COLORS["loss"], bold=True))
        )

    # ── STATISTICI SUMAR ──
    stat_start = last_data_row + 2
    _section_header(ws, stat_start, "STATISTICI PERFORMANȚĂ", 4)

    stats = [
        ("Total Tranzacții Închise", f"=COUNTA(A3:A{last_data_row})"),
        ("Tranzacții Câștigătoare",
         f"=COUNTIF(L3:L{last_data_row},\">0\")"),
        ("Tranzacții Pierzătoare",
         f"=COUNTIF(L3:L{last_data_row},\"<0\")"),
        ("Win Rate (%)",
         f'=IF(COUNTA(A3:A{last_data_row})>0,'
         f'COUNTIF(L3:L{last_data_row},">0")/COUNTA(A3:A{last_data_row}),0)'),
        ("P&L Total Net (RON)", f"=SUM(O3:O{last_data_row})"),
        ("P&L Mediu per Tranzacție",
         f"=IF(COUNTA(A3:A{last_data_row})>0,"
         f"SUM(O3:O{last_data_row})/COUNTA(A3:A{last_data_row}),0)"),
        ("Cel Mai Mare Câștig", f"=MAX(O3:O{last_data_row})"),
        ("Cea Mai Mare Pierdere", f"=MIN(O3:O{last_data_row})"),
        ("Profit Factor",
         f'=IF(ABS(SUMIF(O3:O{last_data_row},"<0",O3:O{last_data_row}))>0,'
         f'SUMIF(O3:O{last_data_row},">0",O3:O{last_data_row})/'
         f'ABS(SUMIF(O3:O{last_data_row},"<0",O3:O{last_data_row})),0)'),
    ]

    for idx, (label, formula) in enumerate(stats):
        r = stat_start + 1 + idx
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = Font(name="Calibri", size=10, bold=True)
        label_cell.border = _border()

        val_cell = ws.cell(row=r, column=2)
        val_cell.value = formula
        val_cell.font = Font(name="Calibri", size=11, bold=True)
        val_cell.border = _border()
        val_cell.alignment = Alignment(horizontal="center")

        if "%" in label and "RON" not in label:
            val_cell.number_format = NUMBER_FORMAT_PERCENT
        elif "RON" in label or "Câștig" in label or "Pierdere" in label or "Mediu" in label:
            val_cell.number_format = NUMBER_FORMAT_CURRENCY
        elif "Factor" in label:
            val_cell.number_format = '0.00'

    # Dropdown pe tip tranzacție
    from openpyxl.worksheet.datavalidation import DataValidation
    formula_dv = '"' + ','.join(TIP_TRANZACTIE) + '"'
    dv = DataValidation(type="list", formula1=formula_dv, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E3:E1000")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

    return ws


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _section_header(ws, row, title, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True,
                     color=COLORS["header_font"])
    cell.fill = PatternFill(start_color=COLORS["title_bg"],
                            end_color=COLORS["title_bg"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
