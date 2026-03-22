"""
Foaia Dashboard — Panou principal portofoliu cu KPI-uri, grafice, alerte
StockAgent — Management Portofoliu Investiții
"""
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from config.stock_config import (
    COLORS, SECTOARE, POZITII_DEMO, PORTFOLIO_INITIAL_CAPITAL,
    PORTFOLIO_CASH, NUMBER_FORMAT_CURRENCY, NUMBER_FORMAT_PERCENT,
    NUMBER_FORMAT_INT,
)


def create_sheet(wb):
    """Creează foaia Dashboard"""
    ws = wb.create_sheet(title="Dashboard")

    # ============================================================
    # TITLU PRINCIPAL
    # ============================================================
    ws.merge_cells("A1:L2")
    title_cell = ws.cell(row=1, column=1,
                         value="STOCKAGENT | DASHBOARD PORTOFOLIU INVESTIȚII")
    title_cell.font = Font(name="Calibri", size=20, bold=True,
                           color=COLORS["header_font"])
    title_cell.fill = PatternFill(start_color=COLORS["title_bg"],
                                  end_color=COLORS["title_bg"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitlu
    ws.merge_cells("A3:L3")
    sub_cell = ws.cell(row=3, column=1,
                       value="Broker Virtual cu 20 de ani experiență | Profil Agresiv | Target 25-40% anual")
    sub_cell.font = Font(name="Calibri", size=11, italic=True,
                         color=COLORS["accent"])
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================
    # KPI-URI (rândul 5-7)
    # ============================================================
    kpi_configs = [
        (1, "Valoare Portofoliu",
         "=SUM('Poziții'!L3:L100)+Configurare!B4",
         COLORS["accent"]),
        (3, "P&L Total (RON)",
         "=SUM('Poziții'!M3:M100)",
         COLORS["profit"]),
        (5, "P&L Total (%)",
         "=IF(SUM('Poziții'!K3:K100)>0,SUM('Poziții'!M3:M100)/SUM('Poziții'!K3:K100),0)",
         COLORS["profit"]),
        (7, "Cash Disponibil",
         "=Configurare!B4",
         COLORS["info"]),
        (9, "Poziții Deschise",
         "=COUNTA('Poziții'!A3:A100)",
         COLORS["neutral"]),
        (11, "Win Rate",
         "=IF(COUNTA('Tranzacții'!A3:A100)>0,"
         "COUNTIF('Tranzacții'!L3:L100,\">0\")/COUNTA('Tranzacții'!A3:A100),0)",
         COLORS["gold"]),
    ]

    for start_col, label, formula, color in kpi_configs:
        _create_kpi_card(ws, start_col, 5, label, formula, color)

    # Formate specifice pe KPI-uri
    ws.cell(row=6, column=1).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=6, column=3).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=6, column=5).number_format = NUMBER_FORMAT_PERCENT
    ws.cell(row=6, column=7).number_format = NUMBER_FORMAT_CURRENCY
    ws.cell(row=6, column=9).number_format = NUMBER_FORMAT_INT
    ws.cell(row=6, column=11).number_format = NUMBER_FORMAT_PERCENT

    # ============================================================
    # ALOCARE SECTORIALĂ (pentru grafic Pie)
    # ============================================================
    ws.cell(row=9, column=1,
            value="ALOCARE PORTOFOLIU PE SECTOR").font = Font(
        name="Calibri", size=12, bold=True, color=COLORS["title_bg"]
    )

    data_start = 10
    ws.cell(row=data_start, column=1, value="Sector").font = Font(bold=True)
    ws.cell(row=data_start, column=2, value="Valoare (RON)").font = Font(bold=True)

    # Calculăm sectoarele unice din pozițiile demo
    sectors_in_portfolio = list(set(p["sector"] for p in POZITII_DEMO))
    sectors_in_portfolio.append("Cash")

    for i, sector in enumerate(sectors_in_portfolio):
        row = data_start + 1 + i
        ws.cell(row=row, column=1, value=sector)
        if sector == "Cash":
            ws.cell(row=row, column=2).value = "=Configurare!B4"
        else:
            ws.cell(row=row, column=2).value = (
                f"=SUMIFS('Poziții'!L3:L100,'Poziții'!E3:E100,\"{sector}\")"
            )
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT_CURRENCY

    # Grafic Pie — Alocare sectorială
    pie = PieChart()
    pie.title = "Alocare Portofoliu pe Sector"
    pie.style = 10
    data_ref = Reference(ws, min_col=2, min_row=data_start,
                         max_row=data_start + len(sectors_in_portfolio))
    cats_ref = Reference(ws, min_col=1, min_row=data_start + 1,
                         max_row=data_start + len(sectors_in_portfolio))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.width = 18
    pie.height = 12
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    ws.add_chart(pie, "D9")

    # ============================================================
    # POZIȚII DESCHISE — REZUMAT RAPID
    # ============================================================
    pos_start = 22
    ws.merge_cells(f"A{pos_start}:C{pos_start}")
    pos_header = ws.cell(row=pos_start, column=1,
                         value="POZIȚII DESCHISE — TOP P&L")
    pos_header.font = Font(name="Calibri", size=12, bold=True,
                           color=COLORS["header_font"])
    pos_header.fill = PatternFill(start_color=COLORS["profit"],
                                  end_color=COLORS["profit"], fill_type="solid")

    # Mini-tabel cu pozitii
    pos_headers = ["Simbol", "P&L (%)", "Status"]
    for c, h in enumerate(pos_headers, 1):
        cell = ws.cell(row=pos_start + 1, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True,
                         color=COLORS["header_font"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"],
                                end_color=COLORS["header_bg"], fill_type="solid")
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")

    for i in range(8):  # Max 8 poziții afișate
        r = pos_start + 2 + i
        data_row = 3 + i
        ws.cell(row=r, column=1).value = f"=IF('Poziții'!B{data_row}<>\"\",'Poziții'!B{data_row},\"\")"
        ws.cell(row=r, column=2).value = f"=IF(A{r}<>\"\",'Poziții'!N{data_row},\"\")"
        ws.cell(row=r, column=2).number_format = NUMBER_FORMAT_PERCENT
        ws.cell(row=r, column=3).value = (
            f'=IF(A{r}="","",IF(\'Poziții\'!N{data_row}>0,"PROFIT","PIERDERE"))'
        )
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = _border()
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    # P&L conditional formatting pe mini-tabel
    pos_range_end = pos_start + 2 + 7
    ws.conditional_formatting.add(
        f"B{pos_start+2}:B{pos_range_end}",
        FormulaRule(formula=[f'AND(B{pos_start+2}<>"",B{pos_start+2}>0)'],
                    fill=PatternFill(start_color=COLORS["light_green"],
                                    end_color=COLORS["light_green"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["profit"], bold=True))
    )
    ws.conditional_formatting.add(
        f"B{pos_start+2}:B{pos_range_end}",
        FormulaRule(formula=[f'AND(B{pos_start+2}<>"",B{pos_start+2}<0)'],
                    fill=PatternFill(start_color=COLORS["light_red"],
                                    end_color=COLORS["light_red"],
                                    fill_type="solid"),
                    font=Font(color=COLORS["loss"], bold=True))
    )

    # ============================================================
    # ALERTE & NOTIFICĂRI
    # ============================================================
    alert_start = 22
    ws.merge_cells(f"E{alert_start}:H{alert_start}")
    alert_header = ws.cell(row=alert_start, column=5,
                           value="ALERTE & NOTIFICĂRI")
    alert_header.font = Font(name="Calibri", size=12, bold=True,
                             color=COLORS["header_font"])
    alert_header.fill = PatternFill(start_color=COLORS["danger"],
                                    end_color=COLORS["danger"], fill_type="solid")

    alerts = [
        ("Poziții aproape de Stop-Loss:",
         "=COUNTIF('Poziții'!P3:P100,\"<0.05\")-COUNTIF('Poziții'!P3:P100,\"\")"),
        ("Poziții în pierdere:",
         "=COUNTIF('Poziții'!M3:M100,\"<0\")"),
        ("Poziții la Target 1:",
         "=COUNTIFS('Poziții'!J3:J100,\">=\"&'Poziții'!Q3:Q100,'Poziții'!A3:A100,\"<>\")"),
        ("Cash sub limită (10%):",
         f'=IF(Configurare!B4/(SUM(\'Poziții\'!L3:L100)+Configurare!B4)<{COLORS["profit"][-2:]},"DA","NU")'),
        ("Sector overweight (>30%):",
         "=\"Verifică Risk Management\""),
    ]

    # Simplified alerts without complex cross-sheet formulas that might error
    simple_alerts = [
        ("Poziții deschise total:", "=COUNTA('Poziții'!A3:A100)"),
        ("Poziții pe profit:", "=COUNTIF('Poziții'!M3:M100,\">0\")"),
        ("Poziții pe pierdere:", "=COUNTIF('Poziții'!M3:M100,\"<0\")"),
        ("Tranzacții închise:", "=COUNTA('Tranzacții'!A3:A100)"),
        ("Watchlist active:", "=COUNTA(Watchlist!A3:A100)"),
    ]

    for idx, (label, formula) in enumerate(simple_alerts, alert_start + 1):
        ws.cell(row=idx, column=5, value=label).font = Font(
            name="Calibri", size=10)
        ws.cell(row=idx, column=5).border = Border(
            bottom=Side(style="thin", color=COLORS["border"]))
        val_cell = ws.cell(row=idx, column=8)
        val_cell.value = formula
        val_cell.font = Font(name="Calibri", size=12, bold=True)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.border = Border(
            bottom=Side(style="thin", color=COLORS["border"]))

    # ============================================================
    # SUMAR INVESTIȚII
    # ============================================================
    sum_start = 33
    ws.merge_cells(f"A{sum_start}:C{sum_start}")
    sal_header = ws.cell(row=sum_start, column=1, value="SUMAR INVESTIȚII")
    sal_header.font = Font(name="Calibri", size=12, bold=True,
                           color=COLORS["header_font"])
    sal_header.fill = PatternFill(start_color=COLORS["accent"],
                                  end_color=COLORS["accent"], fill_type="solid")

    sum_items = [
        ("Capital Inițial:", f"={PORTFOLIO_INITIAL_CAPITAL}"),
        ("Valoare Investită:", "=SUM('Poziții'!K3:K100)"),
        ("Valoare Curentă Poziții:", "=SUM('Poziții'!L3:L100)"),
        ("Cash Disponibil:", "=Configurare!B4"),
        ("Valoare Totală Portofoliu:", "=SUM('Poziții'!L3:L100)+Configurare!B4"),
        ("P&L Realizat (închise):", "=SUM('Tranzacții'!O3:O100)"),
        ("P&L Nerealizat (deschise):", "=SUM('Poziții'!M3:M100)"),
    ]

    for idx, (label, formula) in enumerate(sum_items, sum_start + 1):
        ws.cell(row=idx, column=1, value=label).font = Font(
            name="Calibri", size=10)
        val_cell = ws.cell(row=idx, column=3)
        val_cell.value = formula
        val_cell.number_format = NUMBER_FORMAT_CURRENCY
        val_cell.font = Font(name="Calibri", size=10, bold=True)
        val_cell.alignment = Alignment(horizontal="right")

    # ============================================================
    # NAVIGARE RAPIDĂ
    # ============================================================
    ws.merge_cells(f"K{alert_start}:L{alert_start}")
    nav_header = ws.cell(row=alert_start, column=11, value="NAVIGARE RAPIDĂ")
    nav_header.font = Font(name="Calibri", size=12, bold=True,
                           color=COLORS["header_font"])
    nav_header.fill = PatternFill(start_color=COLORS["accent"],
                                  end_color=COLORS["accent"], fill_type="solid")

    nav_items = [
        "Poziții", "Tranzacții", "Analiză Tehnică", "Analiză Fundamentală",
        "Watchlist", "Risk Management", "Configurare"
    ]

    for idx, sheet_name in enumerate(nav_items, alert_start + 1):
        cell = ws.cell(row=idx, column=11,
                       value=f"→ {sheet_name}")
        cell.font = Font(name="Calibri", size=10, color=COLORS["accent"],
                         underline="single")
        cell.hyperlink = f"#'{sheet_name}'!A1"

    # ============================================================
    # CONFIGURARE COLOANE
    # ============================================================
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["H"].width = 14

    ws.sheet_properties.tabColor = COLORS["accent"]

    return ws


def _create_kpi_card(ws, start_col, start_row, label, formula, color):
    """Creează un card KPI"""
    # Label
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col + 1)
    label_cell = ws.cell(row=start_row, column=start_col)
    label_cell.value = label
    label_cell.font = Font(name="Calibri", size=9, bold=True,
                           color=COLORS["header_font"])
    label_cell.fill = PatternFill(start_color=color, end_color=color,
                                  fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

    # Valoare
    ws.merge_cells(start_row=start_row + 1, start_column=start_col,
                   end_row=start_row + 2, end_column=start_col + 1)
    val_cell = ws.cell(row=start_row + 1, column=start_col)
    val_cell.value = formula
    val_cell.font = Font(name="Calibri", size=16, bold=True, color=color)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")


def _border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)
