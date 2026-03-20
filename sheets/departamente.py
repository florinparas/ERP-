"""
Foaia Departamente - Departamente & funcții
"""
from config.hr_config import (
    DEPARTAMENTE_DEMO, FUNCTII_DEMO, NIVEL_FUNCTIE,
    NUMBER_FORMAT_RON, NUMBER_FORMAT_INT
)
from utils.helpers import (
    create_sheet_with_headers, add_dropdown_validation,
    add_status_conditional_formatting
)
from utils.styles import apply_sheet_title, get_header_font, get_header_fill, get_border, get_center_alignment


def create_sheet(wb):
    """Creează foaia Departamente cu două tabele: Departamente și Funcții"""
    ws = wb.create_sheet(title="Departamente")

    # ============================================================
    # TABEL 1: DEPARTAMENTE (coloanele A-F)
    # ============================================================
    dept_headers = [
        ("ID", 8),
        ("Denumire Departament", 25),
        ("Manager", 20),
        ("Locație", 18),
        ("Buget Lunar (RON)", 18),
        ("Nr. Angajați", 14),
    ]
    num_dept_cols = len(dept_headers)

    # Titlu
    apply_sheet_title(ws, "DEPARTAMENTE", num_dept_cols)

    # Headere
    for col_idx, (name, width) in enumerate(dept_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Date demo
    for row_idx, dept in enumerate(DEPARTAMENTE_DEMO, 3):
        ws.cell(row=row_idx, column=1, value=dept["id"]).border = get_border()
        ws.cell(row=row_idx, column=2, value=dept["denumire"]).border = get_border()
        ws.cell(row=row_idx, column=3, value=dept["manager"]).border = get_border()
        ws.cell(row=row_idx, column=4, value=dept["locatie"]).border = get_border()
        buget_cell = ws.cell(row=row_idx, column=5, value=dept["buget"])
        buget_cell.number_format = NUMBER_FORMAT_RON
        buget_cell.border = get_border()

        # Formula COUNTIF - numără angajații din departamentul respectiv
        dept_name_cell = f"B{row_idx}"
        count_cell = ws.cell(
            row=row_idx, column=6,
            value=f"=COUNTIF('Angajați'!K:K,{dept_name_cell})"
        )
        count_cell.number_format = NUMBER_FORMAT_INT
        count_cell.border = get_border()

    # ============================================================
    # TABEL 2: FUNCȚII (coloanele H-M)
    # ============================================================
    func_start_col = 8  # Coloana H
    func_headers = [
        ("ID", 8),
        ("Denumire Funcție", 25),
        ("Departament", 20),
        ("Nivel", 12),
        ("Salariu Min (RON)", 18),
        ("Salariu Max (RON)", 18),
    ]

    # Titlu funcții
    ws.merge_cells(
        start_row=1, start_column=func_start_col,
        end_row=1, end_column=func_start_col + len(func_headers) - 1
    )
    title_cell = ws.cell(row=1, column=func_start_col, value="FUNCȚII / POSTURI")
    from utils.styles import get_title_font_white, get_title_fill
    from openpyxl.styles import Alignment
    title_cell.font = get_title_font_white()
    title_cell.fill = get_title_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headere funcții
    for col_offset, (name, width) in enumerate(func_headers):
        col = func_start_col + col_offset
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_border()
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width

    # Date demo funcții
    for row_idx, func in enumerate(FUNCTII_DEMO, 3):
        ws.cell(row=row_idx, column=func_start_col, value=func["id"]).border = get_border()
        ws.cell(row=row_idx, column=func_start_col + 1, value=func["denumire"]).border = get_border()
        ws.cell(row=row_idx, column=func_start_col + 2, value=func["dept"]).border = get_border()
        ws.cell(row=row_idx, column=func_start_col + 3, value=func["nivel"]).border = get_border()
        min_cell = ws.cell(row=row_idx, column=func_start_col + 4, value=func["sal_min"])
        min_cell.number_format = NUMBER_FORMAT_RON
        min_cell.border = get_border()
        max_cell = ws.cell(row=row_idx, column=func_start_col + 5, value=func["sal_max"])
        max_cell.number_format = NUMBER_FORMAT_RON
        max_cell.border = get_border()

    # Validare dropdown pentru Nivel
    nivel_col_letter = get_column_letter(func_start_col + 3)
    add_dropdown_validation(
        ws, f"{nivel_col_letter}3:{nivel_col_letter}100", NIVEL_FUNCTIE
    )

    # Validare dropdown pentru Departament (referință la lista departamente)
    dept_col_letter = get_column_letter(func_start_col + 2)
    add_dropdown_validation(
        ws, f"{dept_col_letter}3:{dept_col_letter}100",
        "=Departamente!$B$3:$B$50"
    )

    # Freeze panes
    ws.freeze_panes = "A3"

    return ws
