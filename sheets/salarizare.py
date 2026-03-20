"""
Foaia Salarizare - Calcul salarii conform legislației RO 2025
Include: salariu proporțional, ore suplimentare, deducere personală detaliată
"""
from config.hr_config import (
    CAS_RATE, CASS_RATE, TAX_RATE, CAM_RATE, SALARIU_MINIM_BRUT,
    NUMBER_FORMAT_RON, NUMBER_FORMAT_PERCENT, NUMBER_FORMAT_DATE,
    NUMBER_FORMAT_INT, COLORS
)
from utils.helpers import (
    create_sheet_with_headers, add_status_conditional_formatting
)
from utils.styles import get_border
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def create_sheet(wb):
    """Creează foaia Salarizare cu formule calcul conform legislației RO"""
    headers = [
        ("ID", 8),                          # A
        ("ID Angajat", 12),                  # B
        ("Nume Angajat", 20),                # C
        ("Luna", 8),                         # D
        ("An", 8),                           # E
        ("Salariu Brut Bază", 16),           # F
        ("Zile Lucrate", 12),                # G
        ("Zile Totale Lună", 14),            # H
        ("Salariu Proporțional", 18),        # I
        ("Ore Suplimentare (val)", 18),      # J - NOU
        ("Indemnizație CO", 16),             # K - NOU
        ("Total Brut", 16),                  # L - NOU
        ("CAS 25%", 14),                     # M
        ("CASS 10%", 14),                    # N
        ("Nr. Pers. Întreținere", 16),       # O - NOU
        ("Deducere Personală", 16),          # P
        ("Baza Impozabilă", 16),             # Q
        ("Impozit 10%", 14),                 # R
        ("Alte Deduceri", 14),               # S
        ("Tichete Masă", 14),                # T
        ("Total Rețineri", 14),              # U - NOU
        ("Salariu Net", 16),                 # V
        ("CAM 2.25% (angajator)", 18),       # W
        ("Cost Total Angajator", 18),        # X
    ]

    demo_employees = [
        ("S001", "A001", 3, 2025, 25000, 0),
        ("S002", "A002", 3, 2025, 14000, 0),
        ("S003", "A003", 3, 2025, 7500, 2),
        ("S004", "A004", 3, 2025, 9000, 1),
        ("S005", "A005", 3, 2025, 12000, 0),
    ]

    ws = create_sheet_with_headers(
        wb, "Salarizare", "STAT DE PLATĂ - SALARIZARE", headers
    )

    for row_idx, (sal_id, emp_id, luna, an, brut, pers_intretinere) in enumerate(demo_employees, 3):
        r = row_idx

        # A: ID
        ws.cell(row=r, column=1, value=sal_id).border = get_border()

        # B: ID Angajat
        ws.cell(row=r, column=2, value=emp_id).border = get_border()

        # C: Nume Angajat (VLOOKUP)
        ws.cell(row=r, column=3).value = (
            f'=IFERROR(VLOOKUP(B{r},\'Angajați\'!$A:$C,2,0)&" "'
            f'&VLOOKUP(B{r},\'Angajați\'!$A:$C,3,0),"N/A")'
        )
        ws.cell(row=r, column=3).border = get_border()

        # D, E: Luna, An
        ws.cell(row=r, column=4, value=luna).border = get_border()
        ws.cell(row=r, column=5, value=an).border = get_border()

        # F: Salariu Brut Bază
        cell_brut = ws.cell(row=r, column=6, value=brut)
        cell_brut.number_format = NUMBER_FORMAT_RON
        cell_brut.border = get_border()

        # G: Zile Lucrate (din pontaj)
        ws.cell(row=r, column=7).value = (
            f'=IFERROR(SUMPRODUCT((Pontaj!$A$3:$A$200=B{r})'
            f'*(Pontaj!$D$3:$D$200=D{r})'
            f'*(Pontaj!$E$3:$E$200=E{r})'
            f'*Pontaj!$AK$3:$AK$200),0)'
        )
        ws.cell(row=r, column=7).border = get_border()

        # H: Zile Totale Lună
        ws.cell(row=r, column=8).value = (
            f'=NETWORKDAYS(DATE(E{r},D{r},1),'
            f'EOMONTH(DATE(E{r},D{r},1),0))'
        )
        ws.cell(row=r, column=8).border = get_border()

        # I: Salariu Proporțional = Brut * Zile Lucrate / Zile Totale
        ws.cell(row=r, column=9).value = (
            f'=IF(H{r}>0,ROUND(F{r}*G{r}/H{r},2),0)'
        )
        ws.cell(row=r, column=9).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=9).border = get_border()

        # J: Ore Suplimentare (valoare din foaia Ore Suplimentare)
        ws.cell(row=r, column=10).value = (
            f'=IFERROR(SUMPRODUCT((\'Ore Suplimentare\'!$B$3:$B$200=B{r})'
            f'*(MONTH(\'Ore Suplimentare\'!$D$3:$D$200)=D{r})'
            f'*(YEAR(\'Ore Suplimentare\'!$D$3:$D$200)=E{r})'
            f'*(\'Ore Suplimentare\'!$L$3:$L$200="Aprobat")'
            f'*\'Ore Suplimentare\'!$K$3:$K$200),0)'
        )
        ws.cell(row=r, column=10).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=10).border = get_border()

        # K: Indemnizație CO (simplificat: zile CO * salariu zilnic)
        ws.cell(row=r, column=11).value = (
            f'=IFERROR(SUMPRODUCT((Pontaj!$A$3:$A$200=B{r})'
            f'*(Pontaj!$D$3:$D$200=D{r})'
            f'*(Pontaj!$E$3:$E$200=E{r})'
            f'*Pontaj!$AL$3:$AL$200)'
            f'*IF(H{r}>0,F{r}/H{r},0),0)'
        )
        ws.cell(row=r, column=11).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=11).border = get_border()

        # L: Total Brut = Salariu Proporțional + Ore Suplimentare + Indemnizație CO
        ws.cell(row=r, column=12).value = f'=I{r}+J{r}+K{r}'
        ws.cell(row=r, column=12).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=12).border = get_border()

        # M: CAS 25%
        ws.cell(row=r, column=13).value = f'=ROUND(L{r}*{CAS_RATE},2)'
        ws.cell(row=r, column=13).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=13).border = get_border()

        # N: CASS 10%
        ws.cell(row=r, column=14).value = f'=ROUND(L{r}*{CASS_RATE},2)'
        ws.cell(row=r, column=14).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=14).border = get_border()

        # O: Nr. Persoane Întreținere
        ws.cell(row=r, column=15, value=pers_intretinere).border = get_border()

        # P: Deducere Personală (conform legislație RO detaliată)
        # Deducerea se aplică doar dacă venitul brut lunar <= 2 * salariu minim
        # Valoarea depinde de nr persoane în întreținere
        ws.cell(row=r, column=16).value = (
            f'=IF(L{r}>{SALARIU_MINIM_BRUT}*2,0,'
            f'ROUND(IF(O{r}=0,{SALARIU_MINIM_BRUT}*0.075,'
            f'IF(O{r}=1,{SALARIU_MINIM_BRUT}*0.1,'
            f'IF(O{r}=2,{SALARIU_MINIM_BRUT}*0.15,'
            f'IF(O{r}=3,{SALARIU_MINIM_BRUT}*0.2,'
            f'{SALARIU_MINIM_BRUT}*0.25))))*'
            f'(1-(L{r}-{SALARIU_MINIM_BRUT})/({SALARIU_MINIM_BRUT})),0))'
        )
        ws.cell(row=r, column=16).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=16).border = get_border()

        # Q: Baza Impozabilă = Total Brut - CAS - CASS - Deducere Personală
        ws.cell(row=r, column=17).value = f'=MAX(L{r}-M{r}-N{r}-P{r},0)'
        ws.cell(row=r, column=17).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=17).border = get_border()

        # R: Impozit 10%
        ws.cell(row=r, column=18).value = f'=ROUND(Q{r}*{TAX_RATE},2)'
        ws.cell(row=r, column=18).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=18).border = get_border()

        # S: Alte Deduceri (manual)
        ws.cell(row=r, column=19, value=0).border = get_border()
        ws.cell(row=r, column=19).number_format = NUMBER_FORMAT_RON

        # T: Tichete Masă (manual)
        ws.cell(row=r, column=20, value=0).border = get_border()
        ws.cell(row=r, column=20).number_format = NUMBER_FORMAT_RON

        # U: Total Rețineri = CAS + CASS + Impozit + Alte Deduceri
        ws.cell(row=r, column=21).value = f'=M{r}+N{r}+R{r}+S{r}'
        ws.cell(row=r, column=21).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=21).border = get_border()

        # V: Salariu Net = Total Brut - Total Rețineri + Tichete Masă
        ws.cell(row=r, column=22).value = f'=L{r}-U{r}+T{r}'
        ws.cell(row=r, column=22).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=22).border = get_border()

        # W: CAM 2.25% (contribuție angajator)
        ws.cell(row=r, column=23).value = f'=ROUND(L{r}*{CAM_RATE},2)'
        ws.cell(row=r, column=23).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=23).border = get_border()

        # X: Cost Total Angajator = Total Brut + CAM
        ws.cell(row=r, column=24).value = f'=L{r}+W{r}'
        ws.cell(row=r, column=24).number_format = NUMBER_FORMAT_RON
        ws.cell(row=r, column=24).border = get_border()

    # Rând TOTAL
    total_row = 3 + len(demo_employees)
    total_font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).border = get_border()

    # Totaluri pe coloanele numerice
    sum_cols = [6, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24]
    for col in sum_cols:
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = (
            f'=SUM({col_letter}3:{col_letter}{total_row - 1})'
        )
        ws.cell(row=total_row, column=col).number_format = NUMBER_FORMAT_RON
        ws.cell(row=total_row, column=col).font = total_font
        ws.cell(row=total_row, column=col).border = get_border()

    # Formatare condițională - Salariu Net sub minim net
    net_range = "V3:V1000"
    rule = FormulaRule(
        formula=[f'AND(V3<>"",V3<{SALARIU_MINIM_BRUT}*0.585)'],
        fill=PatternFill(start_color=COLORS["light_red"],
                        end_color=COLORS["light_red"], fill_type="solid"),
        font=Font(color="FF0000", bold=True)
    )
    ws.conditional_formatting.add(net_range, rule)

    # Legendă formule
    legend_row = total_row + 2
    ws.cell(row=legend_row, column=1, value="FORMULE CALCUL:").font = total_font
    formulas_info = [
        "CAS = Total Brut × 25% (contribuție pensie angajat)",
        "CASS = Total Brut × 10% (contribuție sănătate angajat)",
        "Total Brut = Salariu Proporțional + Ore Suplimentare + Indemnizație CO",
        "Deducere Personală = funcție de venitul brut și nr. persoane în întreținere",
        "Baza Impozabilă = Total Brut - CAS - CASS - Deducere Personală",
        "Impozit = Baza Impozabilă × 10%",
        "Total Rețineri = CAS + CASS + Impozit + Alte Deduceri",
        "Salariu Net = Total Brut - Total Rețineri + Tichete Masă",
        "CAM = Total Brut × 2.25% (contribuție angajator)",
        "Cost Total Angajator = Total Brut + CAM",
    ]
    for i, formula_text in enumerate(formulas_info, legend_row + 1):
        ws.cell(row=i, column=1, value=formula_text).font = Font(
            name="Calibri", size=9, italic=True, color="666666"
        )

    return ws
