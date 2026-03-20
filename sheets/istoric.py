"""
Foaia Istoric Modificări - Audit trail / jurnal modificări
"""
from datetime import datetime

from config.hr_config import NUMBER_FORMAT_DATE, COLORS
from utils.helpers import create_sheet_with_headers, add_dropdown_validation
from utils.styles import get_border
from openpyxl.styles import Font


TIPURI_MODIFICARE = [
    "Angajare Nouă",
    "Modificare Date Personale",
    "Modificare Contract",
    "Modificare Salariu",
    "Modificare Departament",
    "Modificare Funcție",
    "Promovare",
    "Suspendare",
    "Reactivare",
    "Demisie",
    "Concediere",
    "Expirare Contract",
    "Act Adițional",
    "Altele",
]

ENTITATI = [
    "Angajați",
    "Contracte",
    "Departamente",
    "Pontaj",
    "Concedii",
    "Salarizare",
    "Evaluări",
    "Training",
    "Recrutare",
    "Documente",
]


def create_sheet(wb):
    """Creează foaia Istoric Modificări"""
    headers = [
        ("ID", 8),
        ("Data Modificării", 16),
        ("Ora", 8),
        ("Utilizator", 18),
        ("ID Angajat", 12),
        ("Nume Angajat", 20),
        ("Entitate Modificată", 18),
        ("Tip Modificare", 22),
        ("Câmp Modificat", 20),
        ("Valoare Veche", 25),
        ("Valoare Nouă", 25),
        ("Motivul Modificării", 25),
    ]

    demo_data = [
        [
            "H001", datetime(2025, 1, 15).date(), "09:30", "Admin",
            "A002", None, "Contracte", "Modificare Salariu",
            "Salariu Brut", "12000 RON", "14000 RON",
            "Mărire salarială anuală",
        ],
        [
            "H002", datetime(2025, 2, 1).date(), "10:15", "Admin",
            "A005", None, "Angajați", "Modificare Departament",
            "Departament", "Marketing", "Vânzări",
            "Restructurare organizațională",
        ],
        [
            "H003", datetime(2025, 3, 1).date(), "08:00", "Admin",
            "A003", None, "Angajați", "Modificare Funcție",
            "Funcție", "Specialist HR Junior", "Specialist HR",
            "Promovare",
        ],
        [
            "H004", datetime(2025, 3, 10).date(), "14:00", "Admin",
            "A004", None, "Concedii", "Altele",
            "Status Concediu", "În Așteptare", "Aprobat",
            "Aprobare concediu medical",
        ],
    ]

    ws = create_sheet_with_headers(
        wb, "Istoric", "ISTORIC MODIFICĂRI / AUDIT TRAIL", headers, demo_data
    )

    # Formule VLOOKUP pentru Nume Angajat
    for row in range(3, 3 + len(demo_data)):
        ws.cell(row=row, column=6).value = (
            f'=IFERROR(VLOOKUP(E{row},\'Angajați\'!$A:$C,2,0)&" "'
            f'&VLOOKUP(E{row},\'Angajați\'!$A:$C,3,0),"N/A")'
        )

    # Formatare date
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                cell.number_format = NUMBER_FORMAT_DATE

    # Validări
    add_dropdown_validation(ws, "G3:G1000", ENTITATI)
    add_dropdown_validation(ws, "H3:H1000", TIPURI_MODIFICARE)

    # Nota informativă
    info_row = 3 + len(demo_data) + 2
    ws.cell(row=info_row, column=1,
            value="NOTĂ: Acest jurnal se completează automat prin modulele VBA "
                  "sau manual la fiecare modificare semnificativă.").font = Font(
        name="Calibri", size=9, italic=True, color="666666"
    )

    # Protecție vizuală - font gri pentru a indica read-only
    ws.sheet_properties.tabColor = "757575"

    return ws
