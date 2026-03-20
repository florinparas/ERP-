"""
Foaia Ore Suplimentare - Evidență și calcul ore suplimentare
"""
from datetime import date

from config.hr_config import (
    NUMBER_FORMAT_DATE, NUMBER_FORMAT_RON, NUMBER_FORMAT_INT,
    NUMBER_FORMAT_PERCENT, COLORS
)
from utils.helpers import (
    create_sheet_with_headers, add_dropdown_validation,
    add_status_conditional_formatting
)


TIPURI_COMPENSARE = ["Plată", "Timp Liber"]
STATUS_OS = ["Solicitat", "Aprobat", "Compensat", "Respins"]
SPORURI_OS = ["75%", "100%"]  # 75% zile lucrătoare, 100% weekend/sărbători


def create_sheet(wb):
    """Creează foaia Ore Suplimentare"""
    headers = [
        ("ID", 8),
        ("ID Angajat", 12),
        ("Nume Angajat", 20),
        ("Data", 14),
        ("Ore Suplimentare", 14),
        ("Motiv", 25),
        ("Aprobat De", 18),
        ("Tip Compensare", 16),
        ("Spor (%)", 10),
        ("Salariu Orar Bază (RON)", 18),
        ("Valoare Brută (RON)", 18),
        ("Status", 14),
        ("Observații", 20),
    ]

    demo_data = [
        [
            "OS001", "A002", None, date(2025, 3, 15), 4,
            "Deadline proiect X", "Popescu Ion",
            "Plată", "75%", None, None, "Aprobat", "",
        ],
        [
            "OS002", "A002", None, date(2025, 3, 22), 8,
            "Deployment producție (sâmbătă)", "Popescu Ion",
            "Plată", "100%", None, None, "Aprobat", "Weekend",
        ],
        [
            "OS003", "A004", None, date(2025, 3, 18), 3,
            "Închidere situații financiare", "Popescu Ion",
            "Timp Liber", "75%", None, None, "Solicitat", "",
        ],
        [
            "OS004", "A005", None, date(2025, 3, 20), 5,
            "Eveniment vânzări special", "Popescu Ion",
            "Plată", "75%", None, None, "Aprobat", "",
        ],
    ]

    ws = create_sheet_with_headers(
        wb, "Ore Suplimentare", "EVIDENȚĂ ORE SUPLIMENTARE", headers, demo_data
    )

    # VLOOKUP Nume Angajat
    for row in range(3, 3 + len(demo_data)):
        ws.cell(row=row, column=3).value = (
            f'=IFERROR(VLOOKUP(B{row},\'Angajați\'!$A:$C,2,0)&" "'
            f'&VLOOKUP(B{row},\'Angajați\'!$A:$C,3,0),"N/A")'
        )

    # Formula Salariu Orar Bază = Salariu Brut / (zile lucrătoare * ore/zi)
    for row in range(3, 3 + len(demo_data)):
        # Salariu orar = Salariu Brut din Contract / (22 zile * 8 ore)
        ws.cell(row=row, column=10).value = (
            f'=IFERROR(ROUND(VLOOKUP(B{row},Contracte!$B:$I,8,0)'
            f'/(22*Configurare!$B$10),2),0)'
        )
        ws.cell(row=row, column=10).number_format = NUMBER_FORMAT_RON

    # Formula Valoare Brută = Ore * Salariu Orar * (1 + Spor%)
    for row in range(3, 3 + len(demo_data)):
        ws.cell(row=row, column=11).value = (
            f'=IF(E{row}<>"",ROUND(E{row}*J{row}'
            f'*(1+SUBSTITUTE(I{row},"%","")/100),2),0)'
        )
        ws.cell(row=row, column=11).number_format = NUMBER_FORMAT_RON

    # Formatare date
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                cell.number_format = NUMBER_FORMAT_DATE

    # Rând TOTAL
    total_row = 3 + len(demo_data) + 1
    from openpyxl.styles import Font
    total_font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=5).value = f'=SUM(E3:E{total_row - 2})'
    ws.cell(row=total_row, column=5).font = total_font
    ws.cell(row=total_row, column=11).value = f'=SUM(K3:K{total_row - 2})'
    ws.cell(row=total_row, column=11).font = total_font
    ws.cell(row=total_row, column=11).number_format = NUMBER_FORMAT_RON

    # Validări
    add_dropdown_validation(ws, "H3:H1000", TIPURI_COMPENSARE)
    add_dropdown_validation(ws, "I3:I1000", SPORURI_OS)
    add_dropdown_validation(ws, "L3:L1000", STATUS_OS)

    # Formatare condițională - Status
    add_status_conditional_formatting(
        ws, "L", 3, 1000,
        {
            "Aprobat": COLORS["light_green"],
            "Solicitat": COLORS["light_yellow"],
            "Compensat": COLORS["light_blue"],
            "Respins": COLORS["light_red"],
        }
    )

    # Legendă
    legend_row = total_row + 2
    ws.cell(row=legend_row, column=1,
            value="LEGISLAȚIE: Conform Codului Muncii, orele suplimentare se compensează cu:").font = Font(
        name="Calibri", size=9, italic=True, color="666666")
    ws.cell(row=legend_row + 1, column=1,
            value="  - Spor 75% pentru zile lucrătoare").font = Font(
        name="Calibri", size=9, italic=True, color="666666")
    ws.cell(row=legend_row + 2, column=1,
            value="  - Spor 100% pentru weekend și sărbători legale").font = Font(
        name="Calibri", size=9, italic=True, color="666666")
    ws.cell(row=legend_row + 3, column=1,
            value="  - Sau timp liber corespunzător în următoarele 60 zile").font = Font(
        name="Calibri", size=9, italic=True, color="666666")

    return ws
